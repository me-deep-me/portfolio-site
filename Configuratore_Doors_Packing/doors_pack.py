import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple
import random
import pandas as pd

# =========================
# PARAMETRI (tuning SHD)
# =========================
# Gap cassa porte
GAP_LEN = 200.0
GAP_W_SIMPLE = 200.0
GAP_W_HERM = 350.0

# Profondità porte
DEPTH_GAP = 100.0
DEPTH_NORMAL = 75.0              # 50 porta + 25 EPS
DEPTH_LEADED = 125.0             # 50 porta + 75 EPS (MODIFICATO)
SINGLE_DOOR_MIN_DEPTH = 400.0

# Imbotti
JAMB_SEC_SIMPLE_H = 100.0        # 200x100 => usiamo "altezza unità impilamento" = 100
JAMB_SEC_HERM_H = 200.0          # 200x200 => unità = 200
JAMB_CRATE_MAX_H = 1000.0
JAMB_CRATE_MAX_D = 1000.0
JAMB_CRATE_DEPTH = 600.0         # fissiamo a 600 per semplicità logistica

# Travi
BEAM_SECTION = 300.0             # 300x300
PALLET_MAX_DEPTH = 1200.0
PALLET_PREF_DEPTH = 900.0        # preferenza
PALLET_STEP = 300.0

# Peso
DENSITY = {"HPL": 236.0, "INOX": 8220.0, "VETRO": 2400.0}
RHO_LEAD = 11340.0
RHO_GLASS = 2400.0
TH_PANEL_M = 46.5 / 1000.0
TH_VISION_M = 20.0 / 1000.0
TH_GLASS_DOOR_M = 12.0 / 1000.0  # vetro porta piena (12 mm)


# =========================
# DECODER STRINGA (Excel)
# =========================
DOUBLE_LEAF_CODES = {"M2S", "M2D", "M2", "A2", "A2S", "A2D", "S2S", "S2D"}

# visiva mapping: code -> (w_mm, h_mm, lead_mm_on_vision?)
VISION_MAP = {
    # ---- CIECA ----
    "V00": (0, 0, 0),

    # ---- STR A VISTA (senza piombo) ----
    "V01": (350, 700, 0),
    "V02": (400, 500, 0),
    "V03": (500, 700, 0),

    "V20": (400, 500, 0),
    "V21": (400, 500, 0),
    "V22": (500, 700, 0),
    "V23": (500, 700, 0),

    "V30": (350, 700, 0),
    "V31": (400, 500, 0),
    "V32": (500, 700, 0),
    "V33": (350, 700, 0),
    "V34": (400, 500, 0),
    "V35": (500, 700, 0),

    # ---- 240x300 ----
    "V04": (240, 300, 0),
    "V05": (240, 300, 0),

    "V06": (240, 300, 1),
    "V07": (240, 300, 1),
    "V08": (240, 300, 2),
    "V09": (240, 300, 2),
    "V10": (240, 300, 3),
    "V11": (240, 300, 3),

    # ---- 400x500 ----
    "V12": (400, 500, 0),
    "V13": (400, 500, 0),
    "V24": (400, 500, 0),
    "V25": (400, 500, 0),
    "V26": (400, 500, 0),
    "V27": (400, 500, 0),

    "V14": (400, 500, 1),
    "V15": (400, 500, 1),
    "V16": (400, 500, 2),
    "V17": (400, 500, 2),
    "V18": (400, 500, 3),
    "V19": (400, 500, 3),
}


def parse_code_string(s: str) -> Dict[str, object]:
    """
    Input esempio:
    PS.APS.TE.02.AAPS.A1S.PB1.V15.N.11.1400.2100. ...

    Regole richieste:
    - pos1: PS/PB => scorrevole/battente
    - pos3: TE => tenuta ermetica, T => tenuta semplice
    - pos4: se in set DOUBLE_LEAF_CODES => 2 ante:
        - larghezza foglio = 0.8*W
        - profondità conta come 2 ante
        - output flag verifica altezza cassa
    - pos5: PB0/PB1/PB2/PB3 => mm piombo (0/1/2/3)
    - pos6: Vxx => visiva (dimensioni + piombo visiva secondo mapping)
    - pos11: larghezza
    - pos12: altezza

    Nota: posizioni contate 1-based sui segmenti separati da "."
    """
    parts = (s or "").split(".")
    # safe get 1-based
    def p(i: int) -> str:
        return parts[i-1].strip() if len(parts) >= i else ""

    pos1 = p(1)  # PS / PB
    pos3 = p(3)  # TE / T
    pos6 = p(6)
    pos5 = p(5)

    w_str = p(11)
    h_str = p(12)

    # sliding/battente
    is_sliding = (pos1 == "PS")
    is_swing = (pos1 == "PB")

    # sealing
    is_herm = (pos3 == "TE")
    # else treat as simple if "T" or anything else
    is_simple = not is_herm

    # double leaf
    is_double = (pos6 in DOUBLE_LEAF_CODES)

    # lead SOLO posizione 7 (1-based)
    mm_pb = 0
    pos7 = p(7)
    pos15 = p(15)

    if pos7.startswith("PB"):
        try:
            mm_pb = int(pos7.replace("PB", ""))
        except:
            mm_pb = 0

    # -----------------------
    # VISIONE (ricerca dinamica Vxx)
    # -----------------------
    vcode = None
    for part in parts:
        part = part.strip()
        if part.startswith("V"):
            vcode = part
            break

    if vcode and vcode in VISION_MAP:
        vis_w, vis_h, vis_pb = VISION_MAP[vcode]
    else:
        vis_w, vis_h, vis_pb = (0, 0, 0)

    senza_visiva = (vis_w == 0 and vis_h == 0)


    # dims
    try:
        w = float(w_str.replace(",", "."))
    except:
        w = 0.0
    try:
        h = float(h_str.replace(",", "."))
    except:
        h = 0.0

    # if double leaf: width leaf considered 80% and depth counts as 2 leaves
    note = ""
    depth_units = 1
    eff_w = w
    if is_double and w > 0:
        depth_units = 2
        eff_w = w / 2.0
        note = "DOPPIA ANTA: verificare altezza cassa (non sempre simmetrica)"

    # door lead: piombata if mm_pb > 0
    piombata = (mm_pb > 0)

    # ========================
    # MATERIALE (posizione 15)
    # ========================
    pos15 = p(15)

    mat_raw = pos15.upper().strip()

    if "VETRO" in mat_raw:
        material = "VETRO"
    elif "INOX" in mat_raw:
        material = "INOX"
    else:
        material = "HPL"

    # door type (scorrevole/battente)
    return {
        "scorrevole": is_sliding,
        "battente": is_swing,
        "ermetica": is_herm,
        "mm_piombo": mm_pb,
        "piombata": piombata,
        "senza_visiva": senza_visiva,
        "visiva_w_mm": float(vis_w),
        "visiva_h_mm": float(vis_h),
        # lead on vision is handled inside weight (we store mm_piombo door; vision lead taken from vcode)
        "visiva_pb_mm": float(vis_pb),
        "w_mm_raw": float(w),
        "h_mm_raw": float(h),
        "w_mm_eff": float(eff_w),
        "depth_units": int(depth_units),
        "note": note,
        "sealing_simple": is_simple,
        "materiale": material,
    }


# =========================
# MODELS
# =========================

@dataclass
class Door:
    door_id: str
    ext_id: str = ""
    zone: str = ""  # "Area/Zona" logistico

    # dimensioni foglio (effettive usate dal motore)
    h_mm: float = 0.0
    w_mm: float = 0.0
    w_mm_total: float = 0.0
    w_mm_leaf_small: float = 0.0    

    scorrevole: bool = False
    battente: bool = False
    vetrata: bool = False
    ermetica_no_maniglione: bool = False  # TE
    piombata: bool = False
    senza_visiva: bool = False

    # peso input
    materiale: str = "HPL"
    mm_piombo: float = 0.0
    visiva_w_mm: float = 0.0
    visiva_h_mm: float = 0.0
    visiva_pb_mm: float = 0.0  # piombo visiva da Vxx

    # double leaf handling
    depth_units: int = 1
    note: str = ""

    double_leaf: bool = False
    asymmetric: bool = False
    w_big_leaf: float = 0.0

    # computed
    peso_kg: float = 0.0

    def area_m2(self) -> float:
        return (self.h_mm / 1000.0) * (self.w_mm / 1000.0)

    def vision_area_m2(self) -> float:
        if self.senza_visiva:
            return 0.0
        return (self.visiva_w_mm / 1000.0) * (self.visiva_h_mm / 1000.0)

    def depth_contrib_mm(self) -> float:
        mat = (self.materiale or "HPL").upper().strip()

        if mat == "VETRO":
            base = 100.0  # 50 porta + 50 EPS
        else:
            base = DEPTH_LEADED if self.piombata else DEPTH_NORMAL

        return base * float(max(1, self.depth_units))

    def crate_len_mm(self) -> float:
        return self.h_mm + GAP_LEN

    def crate_h_mm_simple(self) -> float:
        return self.w_mm + GAP_W_SIMPLE

    def crate_h_mm_herm(self) -> float:
        return self.w_mm + GAP_W_HERM

    def compute_weight(self) -> float:
        """
        Peso porta = peso visiva + peso pannello + peso piombo porta + 15
        Con doppia anta: calcoliamo sul foglio effettivo (w_mm già 0.8*W) e poi moltiplichiamo per depth_units.
        """
        # base areas
        a_door = self.area_m2()
        a_vis = self.vision_area_m2()
        # =========================
        # CASO PORTA VETRO PIENO
        # =========================
        mat = (self.materiale or "HPL").upper().strip()

        if mat == "VETRO":
            lead_th_m = (self.mm_piombo / 1000.0)

            # vetro pieno
            w_glass = a_door * TH_GLASS_DOOR_M * RHO_GLASS

            # piombo porta
            w_lead_door = a_door * lead_th_m * RHO_LEAD

            # visiva opzionale (se mai esiste anche su vetro pieno)
            vision_lead_th_m = (self.visiva_pb_mm / 1000.0)
            w_vision = (a_vis * TH_VISION_M * RHO_GLASS) + (a_vis * vision_lead_th_m * RHO_LEAD)

            w_single_leaf = max(0.0, w_glass + w_lead_door + w_vision + 15.0)

            # +10% peso vetro
            w_single_leaf *= 1.07

            # gestione doppia anta
            if self.depth_units == 2:
                self.peso_kg = w_single_leaf * 2
            else:
                self.peso_kg = w_single_leaf

            return self.peso_kg

        # lead thickness
        lead_th_m = (self.mm_piombo / 1000.0)

        # Piombo porta su area porta
        w_lead_door = (a_door * lead_th_m) * RHO_LEAD

        # Visiva: vetro 20mm + piombo visiva (da Vxx) + piombo porta se presente
        vision_lead_th_m = (self.visiva_pb_mm / 1000.0)
        w_vision = (a_vis * TH_VISION_M * RHO_GLASS) + (a_vis * vision_lead_th_m * RHO_LEAD)

        # Pannello: area pannello = ((W + 174) * H) in mm^2 -> m^2
        a_panel = ((self.w_mm + 174.0) * self.h_mm) / 1_000_000.0
        mat = (self.materiale or "HPL").upper().strip()
        if mat not in DENSITY:
            mat = "HPL"
        rho_panel = DENSITY[mat]
        w_panel = (a_panel * TH_PANEL_M * rho_panel) + (17.0 if mat == "INOX" else 0.0)

        w_single_leaf = max(0.0, w_vision + w_panel + w_lead_door + 15.0)

        # ---- GESTIONE DOPPIA ANTA CORRETTA ----
        if self.depth_units == 2:

            # funzione interna per calcolare peso foglia
            def leaf_weight(width_mm, include_vision=False):

                a_leaf = (self.h_mm / 1000.0) * (width_mm / 1000.0)
                a_panel_leaf = ((width_mm + 174.0) * self.h_mm) / 1_000_000.0

                w_panel_leaf = (a_panel_leaf * TH_PANEL_M * rho_panel) \
                            + (17.0 if mat == "INOX" else 0.0)

                w_lead_leaf = (a_leaf * lead_th_m) * RHO_LEAD

                w_vision_leaf = 0.0
                if include_vision:
                    w_vision_leaf = (a_vis * TH_VISION_M * RHO_GLASS) \
                                    + (a_vis * vision_lead_th_m * RHO_LEAD)

                return max(0.0, w_panel_leaf + w_lead_leaf + w_vision_leaf + 15.0)

            if self.w_mm_leaf_small > 0 and self.w_mm_leaf_small != self.w_mm:
                # asimmetrica reale
                w_big = leaf_weight(self.w_mm, include_vision=True)
                w_small = leaf_weight(self.w_mm_leaf_small, include_vision=False)
                self.peso_kg = w_big + w_small

            else:
                # simmetrica
                w_big = leaf_weight(self.w_mm, include_vision=True)
                w_small = leaf_weight(self.w_mm_leaf_small, include_vision=False)
                self.peso_kg = w_big + w_small

        else:
            self.peso_kg = w_single_leaf
        return self.peso_kg


@dataclass
class DoorCrate:
    crate_id: int
    zone: str
    max_weight_kg: float
    max_depth_mm: float
    doors: List[Door] = field(default_factory=list)
    anti_tip: bool = False

    def total_weight(self) -> float:
        doors_weight = sum(d.peso_kg for d in self.doors)
        return doors_weight + self.crate_structure_weight()

    def count(self) -> int:
        return len(self.doors)

    def has_hermetic(self) -> bool:
        return any(d.ermetica_no_maniglione for d in self.doors)

    def length_mm(self) -> float:
        return max((d.crate_len_mm() for d in self.doors), default=0.0)

    def height_mm(self) -> float:
        if self.has_hermetic():
            return max((d.crate_h_mm_herm() for d in self.doors), default=0.0)
        return max((d.crate_h_mm_simple() for d in self.doors), default=0.0)

    def depth_mm(self) -> float:
        if not self.doors:
            return 0.0
        depth = DEPTH_GAP + sum(d.depth_contrib_mm() for d in self.doors)
        if self.count() == 1:
            depth = max(depth, SINGLE_DOOR_MIN_DEPTH)
        return depth

    def volume_m3(self) -> float:
        return (self.length_mm()/1000.0) * (self.height_mm()/1000.0) * (self.depth_mm()/1000.0)

    def can_add(self, d: Door) -> bool:
        # controllo peso
        if (self.total_weight() + d.peso_kg) > (self.max_weight_kg + 1e-9):
            return False

        # controllo profondità
        current_depth = DEPTH_GAP + sum(x.depth_contrib_mm() for x in self.doors)
        new_depth = current_depth + d.depth_contrib_mm()

        # se è la prima porta consideriamo minimo 400
        if self.count() == 0:
            new_depth = max(new_depth, SINGLE_DOOR_MIN_DEPTH)

        if new_depth > self.max_depth_mm:
            return False

        return True


    def add(self, d: Door) -> None:
        self.doors.append(d)

    def remove(self, d: Door) -> None:
        self.doors.remove(d)

    def crate_structure_weight(self) -> float:
        """
        Peso struttura cassa:
        superficie totale 6 lati
        0.0015 kg per cm2
        """
        L = self.length_mm()
        H = self.height_mm()
        P = self.depth_mm()

        # mm2 -> cm2 ( /100 )
        area_mm2 = 2*(L*H + L*P + H*P)
        area_cm2 = area_mm2 / 100.0

        return area_cm2 * 0.0015



@dataclass
class JambCrate:
    crate_id: int
    zone: str
    sealing: str  # SIMPLE / HERM
    length_mm: float
    depth_mm: float
    height_mm: float
    count: int


@dataclass
class BeamPallet:
    pallet_id: int
    zone: str
    depth_mm: float
    length_mm: float
    count: int


# =========================
# PACKING ENGINE (porte)
# =========================

def objective(crates: List[DoorCrate]) -> Tuple[int, float, float]:
    n = len(crates)
    vol = sum(c.volume_m3() for c in crates)
    ls = [c.length_mm() for c in crates]
    hs = [c.height_mm() for c in crates]
    spread = (max(ls)-min(ls) if ls else 0.0) + (max(hs)-min(hs) if hs else 0.0)
    return (n, vol, spread)

def similarity_cost(crate: DoorCrate, d: Door) -> float:
    new_L = max(crate.length_mm(), d.crate_len_mm())
    # height depends on hermetic presence
    will_be_herm = crate.has_hermetic() or d.ermetica_no_maniglione
    if will_be_herm:
        new_H = max(crate.height_mm(), d.crate_h_mm_herm())
    else:
        new_H = max(crate.height_mm(), d.crate_h_mm_simple())

    cur_vol = crate.volume_m3()
    new_depth = DEPTH_GAP + sum(x.depth_contrib_mm() for x in crate.doors) + d.depth_contrib_mm()
    if crate.count() == 0:
        new_depth = max(new_depth, SINGLE_DOOR_MIN_DEPTH)
    new_vol = (new_L/1000)*(new_H/1000)*(new_depth/1000)

    return (new_vol - cur_vol) + 0.000001*(abs(crate.length_mm()-d.crate_len_mm()) + abs(crate.height_mm()-new_H))

def greedy_pack(doors: List[Door], zone: str, max_weight: float, max_depth: float, start_id: int, rnd: random.Random):
    ds = doors[:]
    # ordine "serio": area, peso, profondità
    ds.sort(key=lambda d: (d.area_m2(), d.peso_kg, d.depth_contrib_mm()), reverse=True)

    # randomizzazione leggera (multi-start)
    window = min(10, len(ds))
    for i in range(len(ds)):
        if window > 1 and rnd.random() < 0.25:
            j = min(len(ds)-1, i + rnd.randint(1, window-1))
            ds[i], ds[j] = ds[j], ds[i]

    crates: List[DoorCrate] = []
    cid = start_id

    for d in ds:
        best = None
        best_cost = float("inf")
        for c in crates:
            if c.can_add(d):
                cost = similarity_cost(c, d)
                if cost < best_cost:
                    best_cost = cost
                    best = c
        if best is None:
            cnew = DoorCrate(crate_id=cid, zone=zone, max_weight_kg=max_weight, max_depth_mm=max_depth)
            cid += 1
            cnew.add(d)
            crates.append(cnew)
        else:
            best.add(d)
    return crates, cid

def try_merge_singletons(crates: List[DoorCrate]) -> List[DoorCrate]:
    if len(crates) <= 1:
        return crates
    changed = True
    while changed:
        changed = False
        singles = [c for c in crates if c.count() == 1]
        if not singles:
            break
        for sc in singles:
            d = sc.doors[0]
            targets = sorted([c for c in crates if c is not sc], key=lambda x: (x.count(), x.volume_m3()))
            for t in targets:
                if t.can_add(d):
                    t.add(d)
                    crates.remove(sc)
                    changed = True
                    break
            if changed:
                break
    return crates

def local_search(crates: List[DoorCrate], rnd: random.Random, max_iters: int = 600) -> List[DoorCrate]:
    best = crates
    best_obj = objective(best)

    def cleanup(crs: List[DoorCrate]) -> List[DoorCrate]:
        return [c for c in crs if c.count() > 0]

    for _ in range(max_iters):
        improved = False

        if len(best) < 2:
            break

        if rnd.random() < 0.6:
            c_from = rnd.choice(best)
            if c_from.count() == 0:
                continue
            d = rnd.choice(c_from.doors)
            candidates = [c for c in best if c is not c_from]
            rnd.shuffle(candidates)
            for c_to in candidates:
                if not c_to.can_add(d):
                    continue
                c_from.remove(d)
                c_to.add(d)
                temp = cleanup(best)
                temp = try_merge_singletons(temp)
                obj = objective(temp)
                if obj < best_obj:
                    best = temp
                    best_obj = obj
                    improved = True
                    break
                else:
                    c_to.remove(d)
                    c_from.add(d)
        else:
            c1, c2 = rnd.sample(best, 2)
            if c1.count() == 0 or c2.count() == 0:
                continue
            d1 = rnd.choice(c1.doors)
            d2 = rnd.choice(c2.doors)
            if (c1.total_weight() - d1.peso_kg + d2.peso_kg) > c1.max_weight_kg:
                continue
            if (c2.total_weight() - d2.peso_kg + d1.peso_kg) > c2.max_weight_kg:
                continue
            c1.remove(d1); c2.remove(d2)
            c1.add(d2); c2.add(d1)
            temp = cleanup(best)
            temp = try_merge_singletons(temp)
            obj = objective(temp)
            if obj < best_obj:
                best = temp
                best_obj = obj
                improved = True
            else:
                c1.remove(d2); c2.remove(d1)
                c1.add(d1); c2.add(d2)

        if not improved and rnd.random() < 0.08:
            break

    return best

def solve_doors(doors: List[Door], max_weight: float, max_depth: float, attempts: int, seed: int):
    for d in doors:
        d.compute_weight()

    # group by zone (empty -> NO_ZONE)
    groups: Dict[str, List[Door]] = {}
    for d in doors:
        z = (d.zone or "").strip() or "NO_ZONE"
        groups.setdefault(z, []).append(d)

    rnd = random.Random(seed)
    best_solution: List[DoorCrate] = []
    best_obj = (10**9, 10**9, 10**9)

    for _ in range(max(1, attempts)):
        cid = 1
        crates_all: List[DoorCrate] = []

        zone_keys = list(groups.keys())
        rnd.shuffle(zone_keys)

        for z in zone_keys:
            g = groups[z][:]
            rnd.shuffle(g)
            crates_z, cid = greedy_pack(g, z, max_weight, max_depth, cid, rnd)
            crates_z = try_merge_singletons(crates_z)
            crates_z = local_search(crates_z, rnd=rnd, max_iters=500)
            crates_z = try_merge_singletons(crates_z)
            crates_all.extend(crates_z)

        obj = objective(crates_all)
        if obj < best_obj:
            best_obj = obj
            best_solution = crates_all

    # single door total
    notes = {"objective": best_obj, "attempts": attempts, "seed": seed, "anti_tip": False}
    if len(doors) == 1 and len(best_solution) == 1:
        best_solution[0].anti_tip = True
        notes["anti_tip"] = True

    # renumber sequential
    best_solution = sorted(best_solution, key=lambda c: c.crate_id)
    for i, c in enumerate(best_solution, start=1):
        c.crate_id = i
    return best_solution, notes


# =========================
# PACKING ENGINE (imbotti)
# =========================

def build_jamb_crates(doors: List[Door]) -> List[JambCrate]:

    groups: Dict[str, List[Door]] = {}
    for d in doors:
        z = (d.zone or "").strip() or "NO_ZONE"
        groups.setdefault(z, []).append(d)

    crates: List[JambCrate] = []
    cid = 1

    BASE_DEPTH = 600.0
    BASE_HEIGHT = 600.0
    MAX_DEPTH = 1000.0
    MAX_HEIGHT = 1000.0
    COLUMN_WIDTH = 200.0

    for z, ds in groups.items():

        units = []
        for d in ds:
            units.append(200 if d.ermetica_no_maniglione else 100)

        units.sort(reverse=True)

        L = (max(d.h_mm for d in ds) if ds else 0.0) + GAP_LEN

        current_depth = BASE_DEPTH
        current_height = 0.0
        current_column_height = 0.0
        current_count = 0

        for u in units:

            # Se entra nella colonna corrente
            if current_column_height + u <= MAX_HEIGHT:
                current_column_height += u
                current_count += 1

            else:
                # Colonna piena → provo nuova colonna
                if current_depth + COLUMN_WIDTH <= MAX_DEPTH:
                    current_depth += COLUMN_WIDTH
                    current_column_height = u
                    current_count += 1
                else:
                    # Cassa piena → chiudo
                    crates.append(JambCrate(
                        crate_id=cid,
                        zone=z,
                        sealing="MIX",
                        length_mm=L,
                        depth_mm=current_depth,
                        height_mm=min(MAX_HEIGHT, max(BASE_HEIGHT, current_column_height)),
                        count=current_count
                    ))
                    cid += 1

                    # reset nuova cassa
                    current_depth = BASE_DEPTH
                    current_column_height = u
                    current_count = 1

        # Chiudi ultima
        if current_count > 0:
            crates.append(JambCrate(
                crate_id=cid,
                zone=z,
                sealing="MIX",
                length_mm=L,
                depth_mm=current_depth,
                height_mm=min(MAX_HEIGHT, max(BASE_HEIGHT, current_column_height)),
                count=current_count
            ))
            cid += 1

    return crates


# =========================
# PACKING ENGINE (travi)
# =========================

def build_beam_pallets(doors: List[Door]) -> List[BeamPallet]:
    """
    Travi:
    - sezione 300x300
    - 900 mm => 12 travi (3 profondità x 4 altezza)
    - 1200 mm => 16 travi (4 x 4)
    - altezza non limitante purché <= profondità (quindi 4 max)
    """

    groups: Dict[str, List[Door]] = {}
    for d in doors:
        if not d.scorrevole:
            continue
        z = (d.zone or "").strip() or "NO_ZONE"
        groups.setdefault(z, []).append(d)

    pallets: List[BeamPallet] = []
    pid = 1

    for z, ds in groups.items():
        n = len(ds)
        if n == 0:
            continue

        # calcolo lunghezze travi
        lengths = []

        for d in ds:
            if d.ermetica_no_maniglione:
                L = (2.0 * d.w_mm) + 500.0
            else:
                L = (2.0 * d.w_mm) + 340.0
            lengths.append(L)
        lengths.sort(reverse=True)

        idx = 0
        while idx < n:
            remaining = n - idx

            if remaining <= 12:
                depth = 900.0
                cap = 12
            else:
                depth = 1200.0
                cap = 16

            take = min(cap, remaining)
            chunk = lengths[idx: idx+take]
            idx += take

            pallets.append(BeamPallet(
                pallet_id=pid,
                zone=z,
                depth_mm=depth,
                length_mm=max(chunk),
                count=take
            ))
            pid += 1

    return pallets


# =========================
# UI (Excel-like)
# =========================

def to_float(s: str) -> float:
    """Converte stringa in float con validazione"""
    if s is None:
        return 0.0
    s = str(s).strip()
    if s == "":
        return 0.0
    s = s.replace(",", ".")
    try:
        value = float(s)
        return value
    except ValueError:
        return 0.0

def validate_numeric_input(value: str, field_name: str, min_val: float = 0) -> tuple[float, str]:
    """Valida input numerico e restituisce (valore, errore)"""
    if value is None or value.strip() == "":
        return 0.0, ""
    
    value = value.strip().replace(",", ".")
    try:
        num_value = float(value)
        if num_value < min_val:
            return 0.0, f"{field_name} non può essere negativo (min: {min_val})"
        return num_value, ""
    except ValueError:
        return 0.0, f"{field_name} deve essere un numero valido"

INPUT_COLUMNS = [
    ("ID", "door_id", "entry", 12),
    ("Ext ID", "ext_id", "entry", 10),
    ("Zona", "zone", "entry", 10),

    ("H (mm)", "h_mm", "entry", 8),
    ("W (mm)", "w_mm", "entry", 8),

    ("Scorr", "scorrevole", "check", 6),
    ("Batt", "battente", "check", 6),
    ("Vetr", "vetrata", "check", 6),
    ("TE", "ermetica_no_maniglione", "check", 6),
    ("2 Ante", "double_leaf", "check", 6),
    ("Asimm", "asym", "check", 6),
    ("W Anta Max", "w_big_leaf", "entry", 8),
    ("Piomb", "piombata", "check", 6),
    ("NoVis", "senza_visiva", "check", 6),

    ("Mat", "materiale", "combo", 6),
    ("mmPb", "mm_piombo", "entry", 6),
    ("VisW", "visiva_w_mm", "entry", 6),
    ("VisH", "visiva_h_mm", "entry", 6),
]

OUTPUT_COLUMNS = [
    ("Sup (m2)", "area_m2"),
    ("Peso(kg)", "peso_kg"),
    ("Note", "note"),
    ("Cassa", "crate_id"),
    ("L", "crate_L"),
    ("H", "crate_H"),
    ("P", "crate_P"),
    ("PesoCassa", "crate_W"),
    ("AntiRib", "anti_tip"),
]

class GridTable(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.canvas = tk.Canvas(self, highlightthickness=0)
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.hsb = ttk.Scrollbar(self, orient="horizontal", command=self.canvas.xview)
        self.inner = ttk.Frame(self.canvas)

        self.inner.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.canvas.configure(yscrollcommand=self.vsb.set,
                            xscrollcommand=self.hsb.set)
        # Scroll con rotella mouse
        self.canvas.bind("<Enter>", lambda e: self.canvas.bind_all("<MouseWheel>", self._on_mousewheel))
        self.canvas.bind("<Leave>", lambda e: self.canvas.unbind_all("<MouseWheel>"))

        self.hsb.grid(row=1, column=0, sticky="ew")
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.vsb.grid(row=0, column=1, sticky="ns")
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.rows: List[Dict[str, object]] = []
        self._build_header()

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


    def _build_header(self):
        for j, (label, key, kind, w) in enumerate(INPUT_COLUMNS):
            ttk.Label(self.inner, text=label, style="Hdr.TLabel").grid(row=0, column=j, padx=3, pady=3, sticky="nsew")
        base = len(INPUT_COLUMNS)
        for j, (label, key) in enumerate(OUTPUT_COLUMNS):
            ttk.Label(self.inner, text=label, style="HdrOut.TLabel").grid(row=0, column=base+j, padx=3, pady=3, sticky="nsew")

    def clear(self):
        for child in self.inner.winfo_children():
            child.destroy()
        self.rows.clear()
        self._build_header()

    def add_row(self, preset: Optional[Dict] = None):
        preset = preset or {}
        r = len(self.rows) + 1
        row: Dict[str, object] = {}
        
        # Disabilitato salvataggio stato automatico per performance
        # if hasattr(self.master, 'save_state'):
        #     self.master.save_state()

        for j, (label, key, kind, w) in enumerate(INPUT_COLUMNS):
            if kind == "entry":
                e = ttk.Entry(self.inner, width=w)
                e.grid(row=r, column=j, padx=2, pady=2, sticky="nsew")
                if key in preset and preset[key] is not None:
                    e.insert(0, str(preset[key]))
                row[key] = e
                
                # Validazione real-time per campi numerici
                if key in ["h_mm", "w_mm", "mm_piombo", "visiva_w_mm", "visiva_h_mm", "w_big_leaf"]:
                    def validate_field(event, widget=e, field_key=key):
                        value = widget.get()
                        field_names = {
                            "h_mm": "Altezza", "w_mm": "Larghezza", 
                            "mm_piombo": "Spessore piombo", 
                            "visiva_w_mm": "Larghezza visiva", 
                            "visiva_h_mm": "Altezza visiva",
                            "w_big_leaf": "Larghezza anta maggiore"
                        }
                        field_name = field_names.get(field_key, field_key)
                        min_val = 0.0
                        if field_key == "h_mm" or field_key == "w_mm":
                            min_val = 100.0  # dimensioni minime porte
                        
                        _, error = validate_numeric_input(value, field_name, min_val)
                        if error:
                            widget.configure(foreground="red")
                            # Mostra errore in tooltip o status
                            if hasattr(widget.master, 'master') and hasattr(widget.master.master, 'set_status'):
                                widget.master.master.set_status(error)
                        else:
                            widget.configure(foreground="black")
                        
                        # Disabilitato salvataggio stato automatico per performance
                        # if event.type == "KeyRelease" and hasattr(widget.master, 'master'):
                        #     if hasattr(widget.master.master, 'save_state'):
                        #         widget.master.master.save_state()
                    
                    e.bind("<KeyRelease>", validate_field)
                    e.bind("<FocusOut>", validate_field)
                
                def go_down(event, col=j):
                    current_widget = event.widget

                    # trova indice riga corrente
                    for i, row_dict in enumerate(self.rows):
                        if row_dict.get(INPUT_COLUMNS[col][1]) == current_widget:
                            next_row = i + 1
                            if next_row < len(self.rows):
                                next_widget = self.rows[next_row].get(INPUT_COLUMNS[col][1])
                                if next_widget:
                                    next_widget.focus_set()
                                    try:
                                        next_widget.selection_range(0, "end")
                                    except:
                                        pass
                                    try:
                                        next_widget.icursor("end")
                                    except:
                                        pass
                            break

                    return "break"  # evita comportamento default

                e.bind("<Return>", go_down)
                if key == "w_big_leaf":
                    e.configure(state="disabled")   
            elif kind == "check":
                var = tk.BooleanVar(value=bool(preset.get(key, False)))
                cb = ttk.Checkbutton(self.inner, variable=var)
                cb.grid(row=r, column=j, padx=2, pady=2)
                row[key] = var
            if key == "asym":
                def toggle_big_leaf(*args, row=row):
                    entry = row["w_big_leaf"]
                    if var.get():
                        entry.configure(state="normal")
                    else:
                        entry.delete(0, "end")
                        entry.configure(state="disabled")

                var.trace_add("write", toggle_big_leaf)
            elif kind == "combo":
                c = ttk.Combobox(self.inner, values=["HPL", "INOX", "VETRO"], width=w, state="readonly")
                c.grid(row=r, column=j, padx=2, pady=2, sticky="nsew")
                val = str(preset.get(key, "HPL")).upper().strip() or "HPL"
                if val not in DENSITY:
                    val = "HPL"
                c.set(val)
                row[key] = c

        base = len(INPUT_COLUMNS)
        for j, (label, key) in enumerate(OUTPUT_COLUMNS):
            width = 10 if key not in ("note",) else 28
            e = ttk.Entry(self.inner, width=width, state="readonly")
            e.grid(row=r, column=base+j, padx=2, pady=2, sticky="nsew")
            row[key] = e

        self.rows.append(row)
        

    def set_out(self, idx: int, key: str, value: str):
        e: ttk.Entry = self.rows[idx][key]  # type: ignore
        e.configure(state="normal")
        e.delete(0, "end")
        e.insert(0, value)
        e.configure(state="readonly")

    def read_doors(self) -> List[Door]:
        doors: List[Door] = []
        for row in self.rows:
            did = row["door_id"].get().strip()  # type: ignore
            if did == "":
                continue
            d = Door(
                door_id=did,
                ext_id=row["ext_id"].get().strip(),  # type: ignore
                zone=row["zone"].get().strip(),  # type: ignore
                h_mm=to_float(row["h_mm"].get()),  # type: ignore
                w_mm=to_float(row["w_mm"].get()),  # type: ignore
                scorrevole=bool(row["scorrevole"].get()),  # type: ignore
                battente=bool(row["battente"].get()),  # type: ignore
                vetrata=bool(row["vetrata"].get()),  # type: ignore
                ermetica_no_maniglione=bool(row["ermetica_no_maniglione"].get()),  # type: ignore
                piombata=bool(row["piombata"].get()),  # type: ignore
                senza_visiva=bool(row["senza_visiva"].get()),  # type: ignore
                materiale=row["materiale"].get().strip(),  # type: ignore
                mm_piombo=to_float(row["mm_piombo"].get()),  # type: ignore
                visiva_w_mm=to_float(row["visiva_w_mm"].get()),  # type: ignore
                visiva_h_mm=to_float(row["visiva_h_mm"].get()),  # type: ignore
                visiva_pb_mm=0.0,
                depth_units=1,
                note="",
            )
            doors.append(d)
        return doors


# =========================
# APP
# =========================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Door Packing Optimizer - Packing Porte / Imbotti / Travi")
        self.state("zoomed")  # Windows full screen
        
        # Storia per undo/redo
        self.history = []
        self.history_index = -1
        
        # Setup shortcuts
        self.setup_shortcuts()

        style = ttk.Style()
        
        # Configura tema moderno
        self.configure(bg='#f0f0f0')  # Sfondo grigio chiaro
        
        # Stili personalizzati
        style.configure("Hdr.TLabel", 
                       font=("Segoe UI", 10, "bold"), 
                       background='#f0f0f0',
                       foreground='#2c3e50')
        
        style.configure("HdrOut.TLabel", 
                       font=("Segoe UI", 10, "bold"), 
                       background='#f0f0f0',
                       foreground='#27ae60')
        
        style.configure("TButton", 
                       padding=(12, 8),
                       font=("Segoe UI", 9),
                       background='#3498db',
                       foreground='white')
        
        style.map("TButton",
                 background=[('active', '#2980b9'), ('pressed', '#21618c')])
        
        style.configure("TFrame", background='#f0f0f0')
        style.configure("TLabelframe", 
                       background='#f0f0f0',
                       borderwidth=2,
                       relief="solid")
        style.configure("TLabelframe.Label", 
                       font=("Segoe UI", 10, "bold"),
                       background='#f0f0f0',
                       foreground='#2c3e50')
        
        # Progress bar personalizzata
        style.configure("TProgressbar",
                       background='#3498db',
                       troughcolor='#ecf0f1',
                       borderwidth=0,
                       lightcolor='#3498db',
                       darkcolor='#3498db')
        
        # Entry personalizzati
        style.configure("TEntry",
                       fieldbackground='white',
                       borderwidth=1,
                       relief="solid",
                       font=("Segoe UI", 9))
        
        # Treeview personalizzato
        style.configure("Treeview",
                       background='white',
                       foreground='#2c3e50',
                       fieldbackground='white',
                       borderwidth=1,
                       font=("Segoe UI", 9))
        style.configure("Treeview.Heading",
                       background='#e74c3c',
                       foreground='white',
                       font=("Segoe UI", 10, "bold"))
        
        # Stile aggiuntivo per intestazioni a contrasto
        style.configure("Results.Treeview.Heading",
                       background='#e74c3c',
                       foreground='#2c3e50',
                       font=("Segoe UI", 10, "bold"),
                       relief="raised")

        # Header principale con titoli
        header_frame = ttk.Frame(self, style="TFrame")
        header_frame.pack(fill="x", padx=20, pady=(15, 10))
        
        title_label = tk.Label(header_frame, 
                              text="🏭 DOOR PACKING OPTIMIZER", 
                              font=("Segoe UI", 16, "bold"),
                              bg='#f0f0f0', fg='#2c3e50')
        title_label.pack(side="left")
        
        subtitle_label = tk.Label(header_frame, 
                                 text="Porte / Imbotti / Travi", 
                                 font=("Segoe UI", 11),
                                 bg='#f0f0f0', fg='#7f8c8d')
        subtitle_label.pack(side="left", padx=(10, 0))
        
        # Progress bar in alto a destra
        self.progress_var = tk.DoubleVar()
        progress_frame = ttk.Frame(header_frame, style="TFrame")
        progress_frame.pack(side="right", padx=(20, 0))
        progress_frame.pack_forget()  # Nascosta di default
        
        self.progress_bar = ttk.Progressbar(
            progress_frame, 
            variable=self.progress_var, 
            maximum=100, 
            length=200,
            mode='determinate',
            style="TProgressbar"
        )
        self.progress_bar.pack(side="left", padx=(0, 8))
        
        self.progress_label = tk.Label(progress_frame, 
                                       text="0%", 
                                       font=("Segoe UI", 8),
                                       bg='#f0f0f0', fg='#7f8c8d')
        self.progress_label.pack(side="left")
        
        self.progress_frame = progress_frame

        # Barra strumenti
        toolbar = ttk.Frame(self, style="TFrame")
        toolbar.pack(fill="x", padx=20, pady=(0, 15))
        
        # Pulsanti principali con icone simulate
        self.create_styled_button(toolbar, "📁 Carica Excel (Elenco Porte)", self.load_excel_decoder, "#27ae60")
        self.create_styled_button(toolbar, "➕ Aggiungi Riga", lambda: self.table.add_row(), "#3498db")
        self.create_styled_button(toolbar, "⚡ Calcola", self.run_all, "#e74c3c")
        self.create_styled_button(toolbar, "💾 Esporta", self.export_excel, "#f39c12")
        
        # Separatore
        ttk.Separator(toolbar, orient='vertical').pack(side="left", padx=15, fill="y")
        
        # Visive globali
        tk.Label(toolbar, text="VisW:", font=("Segoe UI", 8), bg='#f0f0f0', fg='#2c3e50').pack(side="left", padx=(4, 2))
        self.global_vis_w = tk.StringVar()
        ttk.Entry(toolbar, textvariable=self.global_vis_w, width=5).pack(side="left", padx=(0, 5))

        tk.Label(toolbar, text="VisH:", font=("Segoe UI", 8), bg='#f0f0f0', fg='#2c3e50').pack(side="left", padx=(6, 2))
        self.global_vis_h = tk.StringVar()
        ttk.Entry(toolbar, textvariable=self.global_vis_h, width=5).pack(side="left", padx=(0, 15))
        
        # Pulsante applica visive dopo VisH
        self.create_styled_button(toolbar, "👁️ Applica a tutte le visive", self.apply_global_vision, "#9b59b6", smaller=True)
        
        # Separatore verticale dopo le visive
        ttk.Separator(toolbar, orient='vertical').pack(side="left", padx=10, fill="y")

        # Altri parametri nella toolbar
        tk.Label(toolbar, text="Peso max cassa:", font=("Segoe UI", 8), bg='#f0f0f0', fg='#2c3e50').pack(side="left", padx=(5, 2))
        self.max_weight_var = tk.DoubleVar(value=1200.0)
        ttk.Entry(toolbar, textvariable=self.max_weight_var, width=8).pack(side="left", padx=(0, 10))

        tk.Label(toolbar, text="Prof max cassa:", font=("Segoe UI", 8), bg='#f0f0f0', fg='#2c3e50').pack(side="left", padx=(5, 2))
        self.max_depth_var = tk.DoubleVar(value=1100.0)
        ttk.Entry(toolbar, textvariable=self.max_depth_var, width=7).pack(side="left", padx=(0, 10))
        
        tk.Label(toolbar, text="Tent:", font=("Segoe UI", 8), bg='#f0f0f0', fg='#2c3e50').pack(side="left", padx=(5, 2))
        self.attempts_var = tk.IntVar(value=30)
        ttk.Entry(toolbar, textvariable=self.attempts_var, width=4).pack(side="left", padx=(0, 10))

        tk.Label(toolbar, text="Seed:", font=("Segoe UI", 8), bg='#f0f0f0', fg='#2c3e50').pack(side="left", padx=(5, 2))
        self.seed_var = tk.IntVar(value=42)
        ttk.Entry(toolbar, textvariable=self.seed_var, width=5).pack(side="left")

        # Status bar rimossa per pulizia layout

        # Container principale per tabella
        main_container = ttk.Frame(self, style="TFrame")
        main_container.pack(fill="both", expand=True, padx=20, pady=(0, 0))
        
        # Titolo tabella
        table_title = tk.Label(main_container,
                              text="📋 DATI PORTE",
                              font=("Segoe UI", 12, "bold"),
                              bg='#f0f0f0', fg='#2c3e50')
        table_title.pack(anchor="w", pady=(0, 10))

        self.table = GridTable(main_container)
        self.table.pack(fill="both", expand=True)
        

        # =========================
        # DASHBOARD RISULTATI
        # =========================

        # Container dashboard
        dashboard_container = ttk.Frame(self, style="TFrame")
        dashboard_container.pack(fill="x", padx=20, pady=(0, 0))
        
        dashboard_title = tk.Label(dashboard_container,
                                  text="📦 RISULTATI OTTIMIZZAZIONE",
                                  font=("Segoe UI", 11, "bold"),
                                  bg='#f0f0f0', fg='#2c3e50')
        dashboard_title.pack(anchor="w", pady=(0, 0))

        self.results_frame = ttk.Frame(dashboard_container, style="TFrame")
        self.results_frame.pack(fill="x", expand=False)

        pane = ttk.Frame(self.results_frame)
        pane.pack(fill="x", expand=False, padx=10, pady=2)

        # -------- CASSE PORTE --------
        frame_doors = ttk.LabelFrame(pane, text="Casse Porte")
        frame_doors.pack(fill="x", pady=4)

        self.results_doors = ttk.Treeview(
            frame_doors,
            columns=("id", "zona", "n", "L", "H", "P", "peso"),
            show="headings",
            height=5,
            style="Results.Treeview"
        )

        for col, title in zip(
            ("id", "zona", "n", "L", "H", "P", "peso"),
            ("Cassa", "Zona", "N Porte", "L (mm)", "H (mm)", "P (mm)", "Peso (kg)")
        ):
            self.results_doors.heading(col, text=title)
            self.results_doors.column(col, width=110)

        self.results_doors.pack(fill="x")

        # -------- CASSE IMBOTTI --------
        frame_jamb = ttk.LabelFrame(pane, text="Casse Imbotti")
        frame_jamb.pack(fill="x", pady=4)

        self.results_jamb = ttk.Treeview(
            frame_jamb,
            columns=("id", "zona", "n", "L", "H", "P"),
            show="headings",
            height=3,
            style="Results.Treeview"
        )

        for col, title in zip(
            ("id", "zona", "n", "L", "H", "P"),
            ("Cassa", "Zona", "N Imbotti", "L (mm)", "H (mm)", "P (mm)")
        ):
            self.results_jamb.heading(col, text=title)
            self.results_jamb.column(col, width=110)

        self.results_jamb.pack(fill="x")

        # -------- PEDANE TRAVI --------
        frame_beam = ttk.LabelFrame(pane, text="Pedane Travi")
        frame_beam.pack(fill="x", pady=4)

        self.results_beam = ttk.Treeview(
            frame_beam,
            columns=("id", "zona", "n", "prof", "lung"),
            show="headings",
            height=3,
            style="Results.Treeview"
        )

        for col, title in zip(
            ("id", "zona", "n", "prof", "lung"),
            ("Pedana", "Zona", "N Travi", "Prof (mm)", "Lung (mm)")
        ):
            self.results_beam.heading(col, text=title)
            self.results_beam.column(col, width=110)

        self.results_beam.pack(fill="x")

        # Footer migliorato
        footer_frame = ttk.Frame(self, style="TFrame")
        footer_frame.pack(fill="x", side="bottom", padx=20, pady=10)
        
        footer_label = tk.Label(footer_frame, 
                               text="Designed by ME", 
                               font=("Segoe UI", 8, "italic"),
                               bg='#f0f0f0', fg='#95a5a6')
        footer_label.pack(side="left")
        
        # Info shortcuts
        shortcuts_label = tk.Label(footer_frame,
                                  text="Shortcuts: Ctrl+N (Nuova) | Ctrl+O (Apri) | Ctrl+S (Salva) | F5 (Calcola) | Ctrl+Z/Y (Undo/Redo)",
                                  font=("Segoe UI", 8),
                                  bg='#f0f0f0', fg='#bdc3c7')
        shortcuts_label.pack(side="right")


        self.last_doors: List[Door] = []
        self.last_door_crates: List[DoorCrate] = []
        self.last_jamb_crates: List[JambCrate] = []
        self.last_beam_pallets: List[BeamPallet] = []
        self.last_notes: Dict = {}

        self.table.add_row()

    def set_status(self, msg: str):
        # Status bar rimossa - non fare nulla
        pass
    
    def show_progress(self, message: str):
        """Mostra progress bar con messaggio"""
        # Mostra la progress bar nell'header
        self.progress_frame.pack(side="right", padx=(20, 0))
        
        self.progress_var.set(0)
        self.progress_label.config(text="0%")
        self.update_idletasks()
    
    def update_progress(self, value: float, message: str = None):
        """Aggiorna progress bar"""
        self.progress_var.set(value)
        self.progress_label.config(text=f"{int(value)}%")
        self.update_idletasks()
    
    def hide_progress(self):
        """Nasconde progress bar"""
        self.progress_frame.pack_forget()
        self.update_idletasks()
    
    def create_styled_button(self, parent, text, command, color="#3498db", smaller=False):
        """Crea un pulsante stilizzato con colore personalizzato"""
        btn = tk.Button(parent, 
                       text=text,
                       command=command,
                       font=("Segoe UI", 8 if smaller else 9),
                       bg=color,
                       fg='white',
                       relief="flat",
                       cursor="hand2",
                       padx=15 if not smaller else 10,
                       pady=8 if not smaller else 6,
                       borderwidth=0)
        btn.pack(side="left", padx=(0, 8))
        
        # Effetti hover
        def on_enter(e):
            btn.config(bg=self.darken_color(color))
        
        def on_leave(e):
            btn.config(bg=color)
        
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        
        return btn
    
    def darken_color(self, hex_color):
        """Scurisce un colore esadecimale per effetto hover"""
        # Converte hex a RGB
        hex_color = hex_color.lstrip('#')
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        # Scurisce del 20%
        r, g, b = max(0, r - 40), max(0, g - 40), max(0, b - 40)
        return f'#{r:02x}{g:02x}{b:02x}'
    
    def setup_shortcuts(self):
        """Configura shortcuts tastiera"""
        self.bind('<Control-n>', lambda e: self.table.add_row())
        self.bind('<Control-o>', lambda e: self.load_excel_decoder())
        self.bind('<Control-s>', lambda e: self.export_excel())
        self.bind('<Control-Return>', lambda e: self.run_all())
        self.bind('<F5>', lambda e: self.run_all())
        self.bind('<Control-z>', lambda e: self.undo())
        self.bind('<Control-y>', lambda e: self.redo())
    
    def save_state(self):
        """Salva stato corrente per undo/redo"""
        # Salva solo i dati essenziali della tabella
        state = []
        for row in self.table.rows:
            row_data = {}
            for key, widget in row.items():
                if hasattr(widget, 'get'):
                    row_data[key] = widget.get()
                elif hasattr(widget, 'get'):
                    row_data[key] = str(widget.get())
            state.append(row_data)
        
        # Rimuovi storia futura se siamo in mezzo
        if self.history_index < len(self.history) - 1:
            self.history = self.history[:self.history_index + 1]
        
        self.history.append(state)
        self.history_index += 1
        
        # Limita storia a 50 stati
        if len(self.history) > 50:
            self.history = self.history[-50:]
            self.history_index = len(self.history) - 1
    
    def undo(self):
        """Undo ultima modifica"""
        if self.history_index > 0:
            self.history_index -= 1
            self.restore_state(self.history[self.history_index])
            self.set_status("Undo")
    
    def redo(self):
        """Redo modifica"""
        if self.history_index < len(self.history) - 1:
            self.history_index += 1
            self.restore_state(self.history[self.history_index])
            self.set_status("Redo")
    
    def restore_state(self, state):
        """Ripristina stato salvato"""
        self.table.clear()
        for row_data in state:
            self.table.add_row(preset=row_data)

    def apply_global_vision(self):
        w = to_float(self.global_vis_w.get())
        h = to_float(self.global_vis_h.get())

        if w <= 0 or h <= 0:
            messagebox.showwarning("Attenzione", "Inserisci VisW e VisH validi.")
            return

        for row in self.table.rows:
            row["visiva_w_mm"].delete(0, "end")
            row["visiva_w_mm"].insert(0, str(w))
            row["visiva_h_mm"].delete(0, "end")
            row["visiva_h_mm"].insert(0, str(h))

    # -------------------------
    # Import Excel "decoder"
    # -------------------------
    def load_excel_decoder(self):
        """
        Excel:
        - parte da riga 2
        - col A: id porta
        - col B: stringa da decodificare
        """
        path = filedialog.askopenfilename(title="Seleziona Excel (decoder)", filetypes=[("Excel", "*.xlsx *.xls")])
        if not path:
            return
        try:
            df = pd.read_excel(path, header=0)
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile leggere Excel:\n{e}")
            return

        # we assume column A and B by position
        if df.shape[1] < 2:
            messagebox.showerror("Errore", "Excel decoder: servono almeno 2 colonne (A=id, B=stringa).")
            return

        self.table.clear()

        # start from row 2 => dataframe already reads from first data row; if there is header row it's ok.
        # We'll just iterate all rows and skip missing IDs.
        for _, r in df.iterrows():
            did = str(r.iloc[0]).strip()
            code = str(r.iloc[1]).strip()
            if did == "" or did.lower() == "nan":
                continue

            parsed = parse_code_string(code)

            # Door type:
            scorr = bool(parsed["scorrevole"])
            batt = bool(parsed["battente"])
            herm = bool(parsed["ermetica"])
            mm_pb = float(parsed["mm_piombo"])
            piomb = bool(parsed["piombata"])
            novis = bool(parsed["senza_visiva"])
            vis_w = float(parsed["visiva_w_mm"])
            vis_h = float(parsed["visiva_h_mm"])
            vis_pb = float(parsed["visiva_pb_mm"])
            w_eff = float(parsed["w_mm_eff"])
            h_raw = float(parsed["h_mm_raw"])
            depth_units = int(parsed["depth_units"])
            note = str(parsed["note"])

            # By default: materiale HPL (non specificato nel codice stringa)
            preset = {
                "door_id": did,
                "ext_id": "",
                "zone": "NO_ZONE",  # puoi poi editarla
                "h_mm": h_raw,
                "w_mm": w_eff,
                "scorrevole": scorr,
                "battente": batt,
                "vetrata": False,
                "ermetica_no_maniglione": herm,
                "piombata": piomb,
                "senza_visiva": novis,
                "materiale": parsed.get("materiale", "HPL"),
                "mm_piombo": mm_pb,
                "visiva_w_mm": vis_w,
                "visiva_h_mm": vis_h,
            }
            self.table.add_row(preset=preset)

            # scrivo la note in output subito (verrà riscritta anche al run)
            idx = len(self.table.rows) - 1
            self.table.set_out(idx, "note", note)

            # salviamo anche visiva_pb e depth_units e note in un dizionario appoggiato:
            # (qui, per semplicità, li ricostruiamo durante run usando un pass di decode? no: li memorizziamo in ext_id se serve?)
            # migliore: li manteniamo in un dict interno:
            row = self.table.rows[idx]
            row["_decoder_depth_units"] = depth_units
            row["_decoder_vis_pb"] = vis_pb
            row["_decoder_note"] = note

        self.set_status(f"Caricate {len(self.table.rows)} righe (Excel decoder). NB: Zona default NO_ZONE.")

    # -------------------------
    # Run
    # -------------------------
    def run_all(self):
        # read doors from table
        doors: List[Door] = []
        for i, row in enumerate(self.table.rows):
            did = row["door_id"].get().strip()  # type: ignore
            if did == "":
                continue
            d = Door(
                door_id=did,
                ext_id=row["ext_id"].get().strip(),  # type: ignore
                zone=row["zone"].get().strip(),  # type: ignore
                h_mm=to_float(row["h_mm"].get()),  # type: ignore
                w_mm=to_float(row["w_mm"].get()),  # type: ignore
                scorrevole=bool(row["scorrevole"].get()),  # type: ignore
                battente=bool(row["battente"].get()),  # type: ignore
                vetrata=bool(row["vetrata"].get()) if "vetrata" in row else False,  # type: ignore
                ermetica_no_maniglione=bool(row["ermetica_no_maniglione"].get()),  # type: ignore
                piombata=bool(row["piombata"].get()),  # type: ignore
                senza_visiva=bool(row["senza_visiva"].get()),  # type: ignore
                materiale=row["materiale"].get().strip(),  # type: ignore
                mm_piombo=to_float(row["mm_piombo"].get()),  # type: ignore
                visiva_w_mm=to_float(row["visiva_w_mm"].get()),  # type: ignore
                visiva_h_mm=to_float(row["visiva_h_mm"].get()),  # type: ignore
                double_leaf=bool(row["double_leaf"].get()),
                asymmetric=bool(row["asym"].get()),
                w_big_leaf=to_float(row["w_big_leaf"].get()),
                visiva_pb_mm=0.0,
                depth_units=1,
                note="",
            )

            # se arriva dal decoder e depth_units già impostato
            if "_decoder_depth_units" in row:
                d.w_mm_total = to_float(row["w_mm"].get()) * d.depth_units
            else:
                d.w_mm_total = d.w_mm

            # gestione doppia anta
            if d.double_leaf:

                d.depth_units = 2

                if d.asymmetric and d.w_big_leaf > 0:

                    w_big = d.w_big_leaf
                    w_small = d.w_mm_total - w_big

                    if w_small <= 0:
                        messagebox.showerror("Errore", f"Larghezza anta maggiore non valida per {d.door_id}")
                        return

                    d.w_mm = w_big                 # foglia principale
                    d.w_mm_leaf_small = w_small    # foglia secondaria
                    d.note = "DOPPIA ANTA ASIMMETRICA"

                else:
                    # simmetrica
                    d.w_mm = d.w_mm_total / 2.0
                    d.w_mm_leaf_small = d.w_mm
                    d.note = "DOPPIA ANTA SIMMETRICA"

            # apply decoder extras if present
            if "_decoder_depth_units" in row:
                d.depth_units = int(row["_decoder_depth_units"])
            if "_decoder_vis_pb" in row:
                d.visiva_pb_mm = float(row["_decoder_vis_pb"])
            if "_decoder_note" in row:
                d.note = str(row["_decoder_note"])

            doors.append(d)

        if not doors:
            messagebox.showwarning("Dati mancanti", 
                "Nessuna porta da elaborare.\n\n"
                "Per favore:\n"
                "• Inserire almeno un ID porta\n"
                "• Compilare le dimensioni (H, W)\n"
                "• Caricare un file Excel o inserire dati manualmente")
            return

        # Validazione parametri di calcolo
        try:
            max_w = float(self.max_weight_var.get() or 1200.0)
            max_depth = float(self.max_depth_var.get() or 1100.0)
            attempts = int(self.attempts_var.get() or 30)
            seed = int(self.seed_var.get() or 42)
            
            if max_w <= 0:
                messagebox.showwarning("Parametro non valido", "Il peso massimo cassa deve essere maggiore di 0 kg.")
                return
            if max_depth <= 0:
                messagebox.showwarning("Parametro non valido", "La profondità massima cassa deve essere maggiore di 0 mm.")
                return
            if attempts < 1:
                messagebox.showwarning("Parametro non valido", "Il numero di tentativi deve essere almeno 1.")
                return
        except ValueError:
            messagebox.showerror("Errore parametri", "Controllare i valori inseriti nei campi di calcolo.\nAssicurarsi che siano numeri validi.")
            return

        self.show_progress("Calcolo ottimizzazione casse porte...")
        self.update_progress(10, "Analisi porte in corso...")
        
        try:
            door_crates, notes = solve_doors(doors, max_w, max_depth, attempts, seed)
            self.update_progress(50, "Calcolo imbotti...")
            
            jamb_crates = build_jamb_crates(doors)
            self.update_progress(75, "Calcolo pedane travi...")
            
            beam_pallets = build_beam_pallets(doors)
            self.update_progress(90, "Finalizzazione risultati...")
        except Exception as e:
            self.hide_progress()
            messagebox.showerror("Errore durante il calcolo", 
                f"Si è verificato un errore durante l'elaborazione:\n\n{str(e)}\n\n"
                f"Controllare i dati inseriti e riprovare.")
            return

        # map door_id -> crate
        m: Dict[str, DoorCrate] = {}
        for c in door_crates:
            for d in c.doors:
                m[d.door_id] = c

        # write outputs to table
        for i, row in enumerate(self.table.rows):
            did = row["door_id"].get().strip()  # type: ignore
            if did == "" or did not in m:
                continue
            d = next(x for x in doors if x.door_id == did)
            c = m[did]

            self.table.set_out(i, "area_m2", f"{d.area_m2():.3f}")
            self.table.set_out(i, "peso_kg", f"{d.peso_kg:.2f}")
            self.table.set_out(i, "note", d.note or "")
            self.table.set_out(i, "crate_id", str(c.crate_id))
            self.table.set_out(i, "crate_L", f"{c.length_mm():.0f}")
            self.table.set_out(i, "crate_H", f"{c.height_mm():.0f}")
            self.table.set_out(i, "crate_P", f"{c.depth_mm():.0f}")
            self.table.set_out(i, "crate_W", f"{c.total_weight():.2f}")
            self.table.set_out(i, "anti_tip", "SI" if c.anti_tip else "")

        
        self.last_doors = doors
        self.last_door_crates = door_crates
        self.last_jamb_crates = jamb_crates
        self.last_beam_pallets = beam_pallets
        self.last_notes = notes

        self.update_progress(100, "Calcolo completato!")
        self.after(1000, self.hide_progress)  # Nasconde dopo 1 secondo
        self.set_status("Completato.")

        # ===============================
        # AGGIORNA TABELLE RISULTATI
        # ===============================

        # Pulisci
        for t in (self.results_doors, self.results_jamb, self.results_beam):
            for r in t.get_children():
                t.delete(r)

        # --- CASSE PORTE ---
        for c in door_crates:
            self.results_doors.insert(
                "",
                "end",
                values=(
                    c.crate_id,
                    c.zone,
                    c.count(),
                    round(c.length_mm(), 0),
                    round(c.height_mm(), 0),
                    round(c.depth_mm(), 0),
                    round(c.total_weight(), 2)
                )
            )

        # --- CASSE IMBOTTI ---
        for jc in jamb_crates:
            self.results_jamb.insert(
                "",
                "end",
                values=(
                    jc.crate_id,
                    jc.zone,
                    jc.count,
                    round(jc.length_mm, 0),
                    round(jc.height_mm, 0),
                    round(jc.depth_mm, 0),
                )
            )

        # --- PEDANE TRAVI ---
        for bp in beam_pallets:
            self.results_beam.insert(
                "",
                "end",
                values=(
                    bp.pallet_id,
                    bp.zone,
                    bp.count,
                    round(bp.depth_mm, 0),
                    round(bp.length_mm, 0),
                )
            )



    # -------------------------
    # Export
    # -------------------------
    def export_excel(self):

        if not self.last_doors:
            messagebox.showwarning("Attenzione", "Prima esegui 'Calcola & Ottimizza'.")
            return

        path = filedialog.asksaveasfilename(
            title="Salva Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return

        from openpyxl import Workbook
        from openpyxl.styles import Font, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        # =====================================================
        # FOGLIO 1 — LAYOUT VISIVO
        # =====================================================

        ws = wb.active
        ws.title = "LAYOUT_VISIVO"

        from openpyxl.styles import Font, Border, Side
        from openpyxl.utils import get_column_letter

        thick = Side(border_style="medium")
        border_thick = Border(bottom=thick)

        row = 1

        ws.cell(row=row, column=1).value = "CASSE PORTE"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 2

        # Mappa porte → cassa
        crate_map = {}
        for c in self.last_door_crates:
            for d in c.doors:
                crate_map.setdefault(c.crate_id, []).append(d)

        # -------- CASSE PORTE --------
        for cid in sorted(crate_map.keys()):
            doors = crate_map[cid]
            c = self.last_door_crates[cid-1]

            for d in doors:
                ws.cell(row=row, column=1).value = d.door_id
                ws.cell(row=row, column=2).value = d.ext_id   # <-- NUOVO
                ws.cell(row=row, column=3).value = d.h_mm
                ws.cell(row=row, column=4).value = d.w_mm
                ws.cell(row=row, column=5).value = "S" if d.scorrevole else "B"
                ws.cell(row=row, column=6).value = d.mm_piombo
                row += 1

            # Riga separazione
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = border_thick
            row += 1

            ws.cell(row=row, column=8).value = (
                f"CASSA {c.crate_id} | ZONA {c.zone} | "
                f"{int(c.length_mm())}x"
                f"{int(c.depth_mm())}x"
                f"{int(c.height_mm())}h"
            )
            ws.cell(row=row, column=8).font = Font(bold=True)
            row += 2

        # =====================================================
        # SEZIONE IMBOTTI
        # =====================================================

        ws.cell(row=row, column=1).value = "CASSE IMBOTTI"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 2

        for jc in self.last_jamb_crates:
            ws.cell(row=row, column=1).value = f"Zona: {jc.zone}"
            ws.cell(row=row, column=2).value = f"Imbotti: {jc.count}"
            row += 1

            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border_thick
            row += 1

            ws.cell(row=row, column=8).value = (
                f"cassa imbotti {int(jc.length_mm)}x"
                f"{int(jc.depth_mm)}x"
                f"{int(jc.height_mm)}h"
            )
            ws.cell(row=row, column=8).font = Font(bold=True)
            row += 2

        # =====================================================
        # SEZIONE PEDANE TRAVI
        # =====================================================

        ws.cell(row=row, column=1).value = "PEDANE TRAVI"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 2

        for bp in self.last_beam_pallets:
            ws.cell(row=row, column=1).value = f"Zona: {bp.zone}"
            ws.cell(row=row, column=2).value = f"Travi: {bp.count}"
            row += 1

            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border_thick
            row += 1

            ws.cell(row=row, column=8).value = (
                f"pedana travi {int(bp.length_mm)}x{int(bp.depth_mm)}"
            )
            ws.cell(row=row, column=8).font = Font(bold=True)
            row += 2

        # Auto width
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15


        # =====================================================
        # FOGLIO 2 — ELENCO PORTE
        # =====================================================

        ws2 = wb.create_sheet("ELENCO_PORTE")

        headers = [
            "id", "ext_id", "zona", "h_mm",
            "w_totale", "w_foglia_princ", "w_foglia_sec",
            "scorrevole", "battente",
            "piombata", "mm_piombo",
            "peso_kg", "cassa_id"
        ]

        for col, h in enumerate(headers, 1):
            ws2.cell(row=1, column=col).value = h
            ws2.cell(row=1, column=col).font = Font(bold=True)

        row = 2

        for d in self.last_doors:
            c = crate_map.get(d.door_id)
            ws2.cell(row=row, column=1).value = d.door_id
            ws2.cell(row=row, column=2).value = d.ext_id
            ws2.cell(row=row, column=3).value = d.zone
            ws2.cell(row=row, column=4).value = d.h_mm

            ws2.cell(row=row, column=5).value = d.w_mm_total
            ws2.cell(row=row, column=6).value = d.w_mm
            ws2.cell(row=row, column=7).value = d.w_mm_leaf_small

            ws2.cell(row=row, column=8).value = d.scorrevole
            ws2.cell(row=row, column=9).value = d.battente
            ws2.cell(row=row, column=10).value = d.piombata
            ws2.cell(row=row, column=11).value = d.mm_piombo
            ws2.cell(row=row, column=12).value = round(d.peso_kg, 2)
            ws2.cell(row=row, column=13).value = crate_map.get(d.door_id).crate_id if d.door_id in crate_map else ""
            row += 1

        for col in range(1, 14):
            ws2.column_dimensions[get_column_letter(col)].width = 15

        # =====================================================
        # FOGLIO 3 — RIEPILOGO CASSE
        # =====================================================

        ws3 = wb.create_sheet("RIEPILOGO_CASSE")

        headers = ["Cassa", "Zona", "N Porte", "L", "H", "P", "Peso", "Volume"]
        for col, h in enumerate(headers, 1):
            ws3.cell(row=1, column=col).value = h
            ws3.cell(row=1, column=col).font = Font(bold=True)

        row = 2

        for c in self.last_door_crates:
            ws3.cell(row=row, column=1).value = c.crate_id
            ws3.cell(row=row, column=2).value = c.zone
            ws3.cell(row=row, column=3).value = c.count()
            ws3.cell(row=row, column=4).value = round(c.length_mm(), 0)
            ws3.cell(row=row, column=5).value = round(c.height_mm(), 0)
            ws3.cell(row=row, column=6).value = round(c.depth_mm(), 0)
            ws3.cell(row=row, column=7).value = round(c.total_weight(), 2)
            ws3.cell(row=row, column=8).value = round(c.volume_m3(), 3)
            row += 1

        for jc in self.last_jamb_crates:
            ws3.cell(row=row, column=1).value = f"IMB-{jc.crate_id}"
            ws3.cell(row=row, column=2).value = jc.zone
            ws3.cell(row=row, column=3).value = jc.count
            ws3.cell(row=row, column=4).value = round(jc.length_mm, 0)
            ws3.cell(row=row, column=5).value = round(jc.height_mm, 0)
            ws3.cell(row=row, column=6).value = round(jc.depth_mm, 0)
            ws3.cell(row=row, column=7).value = ""
            ws3.cell(row=row, column=8).value = round(
                (jc.length_mm/1000)*(jc.height_mm/1000)*(jc.depth_mm/1000),
                3
            )
            row += 1    

        for bp in self.last_beam_pallets:
            ws3.cell(row=row, column=1).value = f"TRV-{bp.pallet_id}"
            ws3.cell(row=row, column=2).value = bp.zone
            ws3.cell(row=row, column=3).value = bp.count
            ws3.cell(row=row, column=4).value = round(bp.length_mm, 0)
            ws3.cell(row=row, column=5).value = ""
            ws3.cell(row=row, column=6).value = round(bp.depth_mm, 0)
            ws3.cell(row=row, column=7).value = ""
            ws3.cell(row=row, column=8).value = round(
                (bp.length_mm/1000)*(bp.depth_mm/1000)*0.3,
                3
            )
            row += 1

        for col in range(1, 12):
            ws3.column_dimensions[get_column_letter(col)].width = 15

        wb.save(path)
        messagebox.showinfo("OK", "Excel esportato con layout 3 fogli.")


if __name__ == "__main__":
    App().mainloop()
