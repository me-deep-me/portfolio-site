"""
gui.py

Interfaccia grafica Tkinter per il sistema di nesting.
"""

import sys
import os
import threading
import queue
import logging
import pandas as pd

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, simpledialog
    _HAS_TK = True
except Exception:
    _HAS_TK = False

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    _HAS_DND = True
except Exception:
    _HAS_DND = False

from config import CONCISE_FILTER


# ============================================================================
# LOGGER HANDLERS PER GUI
# ============================================================================

import re
import time as _time


LEVEL_GLYPH = {
    "DEBUG":   "·",
    "INFO":    " ",
    "WARNING": "!",
    "ERROR":   "✖",
    "CRITICAL":"✖",
}


class UserFriendlyPolicy:
    """Trasforma i messaggi tecnici in step user-friendly per la console.

    Restituisce liste di tuple (msg, level) da stampare, o lista vuota per
    scartare il messaggio.
    """
    def __init__(self):
        self.reset()

    def reset(self):
        self._xlsx_announced = False
        self._png_announced = False
        self._3d_announced = False

    # ---- log records (livello già noto) ----
    def transform_log(self, msg, level):
        s = (msg or "").strip()
        if not s:
            return []

        # ----- DROP rumore (solo se INFO; warning/error vengono comunque mostrati a meno di match esplicito) -----
        DROP_INFO = (
            "🚀 Heuristic pack",
            "📦 Pannelli inizialmente posizionati",
            "🆕 Pannello",
            "✅ Pannelli finali allocati",
            "✅ Nessun pannello escluso",
            "🧪 Baseline sheets",
            "🔄 Lettura pannelli",
            "🖨️",
        )
        if level == "INFO" and any(s.startswith(p) for p in DROP_INFO):
            return []

        # WARNING di dettaglio "❗ Pannelli NON allocati: N" e elenchi "- 18 (#-) ..."
        if "Pannelli NON allocati" in s:
            return []
        if s.startswith("- ") and "(#-)" in s:
            return []

        # ----- REWRITES user-friendly -----
        if s.startswith("🔄 Conversione preliminare"):
            return [("📂 Lettura file Excel in corso…", "INFO")]
        if s.startswith("➜ Totale pannelli espansi:"):
            n = s.split(":")[-1].strip()
            return [(f"📋 Pannelli totali da posizionare: {n}", "INFO")]
        if s.startswith("📦 Packing completato:"):
            tail = s.split(":", 1)[1].strip()
            return [(f"✅ Nesting completato — {tail}", "INFO")]
        if s.startswith("🎨 Generazione PNG"):
            if self._png_announced:
                return []
            self._png_announced = True
            return [("🎨 Generazione anteprime PNG dei layout…", "INFO")]
        if s.startswith("✅ Generati") and "PNG" in s:
            new = s.replace("Generati", "Generate").replace("di layout", "anteprime PNG")
            return [(new, "INFO")]
        if s.startswith("📸 PNG esploso 3D – overview"):
            return [("🧊 Generazione viste 3D delle casse di imballo…", "INFO")]
        if s.startswith("📸 PNG esploso 3D – Crate"):
            m = re.search(r"Crate\s+(\d+)", s)
            num = m.group(1) if m else "?"
            return [(f"🧊 Vista 3D cassa #{num} generata", "INFO")]
        if s.startswith("⚠️ GLB non esportato"):
            return []  # nasconde dettaglio tecnico

        # default: passa
        return [(msg, level)]

    # ---- stdout/stderr (livello assente) ----
    def transform_raw(self, line):
        s = (line or "").strip()
        if not s:
            return None  # blank: caller decide

        if s.startswith("👉 Colonne viste da Pandas"):
            return []
        if s.startswith("[") and s.endswith("]") and "'" in s:
            return []  # echo della lista colonne
        if s.startswith("✅ File convertito"):
            return []
        if s.startswith("XLSX: scritto foglio"):
            if self._xlsx_announced:
                return []
            self._xlsx_announced = True
            return [("💾 Scrittura workbook XLSX in corso…", "INFO")]
        if s.startswith("📒 Workbook unico:"):
            path = s.split(":", 1)[1]
            if "(engine" in path:
                path = path.split("(engine")[0]
            return [(f"💾 Workbook salvato: {path.strip()}", "INFO")]

        # default: mostra come RAW
        return [(line, "RAW")]


class TextLoggerHandler(logging.Handler):
    """Handler che redirige i log verso la GUI applicando la policy user-friendly."""
    def __init__(self, write_fn, policy=None):
        super().__init__()
        self.write_fn = write_fn
        self.policy = policy

    def emit(self, record):
        try:
            raw_msg = record.getMessage()
            level = record.levelname
            if self.policy is not None:
                pairs = self.policy.transform_log(raw_msg, level)
            else:
                pairs = [(raw_msg, level)]
            ts = _time.strftime("%H:%M:%S")
            for msg, lvl in pairs:
                glyph = LEVEL_GLYPH.get(lvl, " ")
                self.write_fn(f"{ts}  {glyph}  {msg}\n", lvl)
        except Exception:
            pass


class CompactFormatter(logging.Formatter):
    """Formatter compatto (kept per compatibilità, ora non più usato)."""
    LEVEL_GLYPH = LEVEL_GLYPH

    def format(self, record):
        ts = self.formatTime(record, "%H:%M:%S")
        glyph = LEVEL_GLYPH.get(record.levelname, " ")
        return f"{ts}  {glyph}  {record.getMessage()}"


class StreamToText:
    """Redirect stdout/stderr verso una funzione di scrittura."""
    def __init__(self, write_fn):
        self.write_fn = write_fn
    
    def write(self, buf):
        if buf:
            self.write_fn(buf)
    
    def flush(self):
        pass


# ============================================================================
# NAMESPACE HELPER
# ============================================================================

class _ArgsNamespace:
    """Namespace per passare argomenti alla funzione runner."""
    def __init__(self, **kwargs):
        for k, v in kwargs.items():
            setattr(self, k, v)


# ============================================================================
# GUI APP
# ============================================================================

class NestGUI:
    """Interfaccia grafica principale per il nesting."""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Panel Nesting Optimizer (GUI)")

        # ========= STATE =========
        self.selected_file = tk.StringVar(value="")
        self.sheet_w = tk.StringVar(value="930")
        self.sheet_h = tk.StringVar(value="3000")
        self.kerf = tk.StringVar(value="4")
        self.no_rotate = tk.BooleanVar(value=False)
        self.visualise = tk.BooleanVar(value=False)
        self.group_run = tk.BooleanVar(value=True)
        self.id_order = tk.StringVar(value="asc")
        self.out_prefix = None
        self.max_crate_kg = tk.StringVar(value="1200")
        self.concise_log = tk.BooleanVar(value=True)
        self.sheet_choice = tk.StringVar(value="none")
        self.dark_mode = tk.BooleanVar(value=True)
        self._running = False
        self._progress = None

        # ========= LAYOUT =========
        self._build_ui()
        
        # Coda per thread-safe updates
        self._q = queue.Queue()
        
        # Setup logging
        self._setup_logging()
        
        # Drain loop
        self._drain()
        
        # Applica tema iniziale
        try:
            self._apply_theme()
        except Exception:
            pass

    def _build_ui(self):
        """Costruisce l'interfaccia utente."""
        # Header
        header = ttk.Frame(self.root, style="Header.TFrame")
        header.pack(fill=tk.X)
        hl = ttk.Label(header, text="Æ-NEST ~ Panel Nesting Optimizer", style="Header.TLabel")
        hl.pack(side=tk.LEFT, padx=10, pady=8)
        hs = ttk.Label(header, text="Smart nesting • DXF • XLSX • 3D", style="SubHeader.TLabel")
        hs.pack(side=tk.LEFT, padx=8)

        # Notebook con due pagine: Nesting (esistente) e Output XLSX
        self.nb = ttk.Notebook(self.root, style="Dark.TNotebook")
        self.nb.pack(fill=tk.BOTH, expand=True)

        tab_main = ttk.Frame(self.nb, style="Card.TFrame")
        tab_view = ttk.Frame(self.nb, style="Card.TFrame")
        self.nb.add(tab_main, text="Nesting")
        self.nb.add(tab_view, text="Output XLSX")

        # ===== Tab 1: UI esistente =====
        # Paned verticale: area principale sopra, console sotto (ridimensionabile)
        vertical_paned = ttk.Panedwindow(tab_main, orient=tk.VERTICAL)
        vertical_paned.pack(fill=tk.BOTH, expand=True)
        try:
            vertical_paned.configure(style="Dark.TPanedwindow")
        except Exception:
            pass

        # TOP AREA: contenuto principale (albero + opzioni)
        top_area = ttk.Frame(vertical_paned, style="Card.TFrame")
        vertical_paned.add(top_area, weight=3)

        # Paned orizzontale dentro top_area
        paned = ttk.Panedwindow(top_area, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True)
        try:
            paned.configure(style="Dark.TPanedwindow")
        except Exception:
            pass

        # LEFT: preview
        left = ttk.Frame(paned, style="Card.TFrame")
        paned.add(left, weight=3)
        
        top_file = ttk.Frame(left, style="Card.TFrame")
        top_file.pack(fill=tk.X, padx=6, pady=6)
        ttk.Label(top_file, text="File:").pack(side=tk.LEFT)
        self.ent_file = ttk.Entry(top_file, textvariable=self.selected_file)
        self.ent_file.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=6)
        self.btn_open = ttk.Button(top_file, text="Apri…", command=self.open_file)
        self.btn_open.pack(side=tk.LEFT)
        self.btn_template = ttk.Button(top_file, text="Template XLSX", command=self.download_template)
        self.btn_template.pack(side=tk.LEFT, padx=(6,0))
        
        # Toggle Dark Mode
        ttk.Checkbutton(top_file, text="Dark mode", variable=self.dark_mode, command=self._apply_theme).pack(side=tk.LEFT, padx=(12,0))
        
        # Drag & Drop
        if _HAS_DND:
            self.ent_file.drop_target_register(DND_FILES)
            self.ent_file.dnd_bind('<<Drop>>', self._on_drop)

        # Treeview preview (senza scrollbars visibili)
        self.tree = ttk.Treeview(left, columns=(), show="headings", height=18, style="Dark.Treeview")
        self.tree.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0,6))
        # Mouse wheel scroll (verticale) e Shift+wheel (orizzontale)
        def _on_mousewheel(event):
            delta = int(-1*(event.delta/120)) if hasattr(event, 'delta') else 0
            try:
                self.tree.yview_scroll(delta, 'units')
            except Exception:
                pass
            return "break"
        def _on_shift_wheel(event):
            delta = int(-1*(event.delta/120)) if hasattr(event, 'delta') else 0
            try:
                self.tree.xview_scroll(delta, 'units')
            except Exception:
                pass
            return "break"
        self.tree.bind('<MouseWheel>', _on_mousewheel)
        self.tree.bind('<Shift-MouseWheel>', _on_shift_wheel)

        # RIGHT: options
        right = ttk.Frame(paned, style="Card.TFrame")
        paned.add(right, weight=2)
        frm = ttk.LabelFrame(right, text="Impostazioni", style="Card.TLabelframe")
        frm.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        # Dimensioni lastra
        g1 = ttk.LabelFrame(frm, text="Dimensioni lastra", style="Card.TLabelframe")
        g1.pack(fill=tk.X, padx=4, pady=6)

        self._build_sheet_options(g1)

        # Kerf
        rowK = ttk.Frame(g1)
        rowK.pack(fill=tk.X, padx=4, pady=4)
        ttk.Label(rowK, text="Kerf (mm):").pack(side=tk.LEFT)
        ttk.Entry(rowK, width=6, textvariable=self.kerf).pack(side=tk.LEFT, padx=(6,0))

        # Opzioni generali
        g2 = ttk.Frame(frm, style="Card.TFrame")
        g2.pack(fill=tk.X, padx=4, pady=4)
        ttk.Checkbutton(g2, text="No rotate", variable=self.no_rotate).pack(side=tk.LEFT)
        ttk.Checkbutton(g2, text="Visualise (PNG)", variable=self.visualise).pack(side=tk.LEFT, padx=(12,0))
        ttk.Checkbutton(g2, text="Group run (CNC)", variable=self.group_run).pack(side=tk.LEFT, padx=(12,0))
        ttk.Checkbutton(g2, text="Log sintetico", variable=self.concise_log).pack(side=tk.LEFT, padx=(12,0))

        # ID order
        g3 = ttk.Frame(frm, style="Card.TFrame")
        g3.pack(fill=tk.X, padx=4, pady=4)
        ttk.Label(g3, text="id-order:").pack(side=tk.LEFT)
        ttk.Combobox(g3, width=6, state="readonly", textvariable=self.id_order,
                     values=("asc","desc")).pack(side=tk.LEFT, padx=(6,0))

        # Packaging
        gPack = ttk.LabelFrame(frm, text="Packaging", style="Card.TLabelframe")
        gPack.pack(fill=tk.X, padx=4, pady=6)
        rowP = ttk.Frame(gPack, style="Card.TFrame")
        rowP.pack(fill=tk.X, padx=4, pady=2)
        ttk.Label(rowP, text="Peso max cassa (kg):").pack(side=tk.LEFT)
        ttk.Entry(rowP, width=8, textvariable=self.max_crate_kg).pack(side=tk.LEFT, padx=(6,0))

        # Bottone azione
        actions = ttk.Frame(frm, style="Card.TFrame")
        actions.pack(fill=tk.X, padx=4, pady=8)
        self.btn_start = ttk.Button(actions, text="Inizia il calcolo", command=self.ask_output_name)
        self.btn_start.pack(side=tk.LEFT)

        # CONSOLE (nel paned verticale, ridimensionabile)
        console_frame = ttk.Frame(vertical_paned, style="Card.TFrame")
        vertical_paned.add(console_frame, weight=1)
        ttk.Label(console_frame, text="Console:").pack(anchor="w", padx=6)
        self.console = tk.Text(console_frame, height=12, wrap="word")
        self.console.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0,6))
        self.console.configure(state="disabled")

        # ===== Tab 2: Viewer XLSX =====
        self._build_output_viewer(tab_view)

    def _apply_theme(self):
        """Applica tema chiaro/scuro con accenti neon verde."""
        style = getattr(self, "style", None)
        if style is None:
            style = ttk.Style(self.root)
            self.style = style

        accent  = "#2F81F7"
        bg_dark = "#0F1117"
        bg_panel = "#161B22"
        bg_input = "#21262D"
        fg_text = "#E6EDF3"
        fg_muted = "#8B949E"
        sel_bg  = "#1F3448"
        border  = "#30363D"

        if self.dark_mode.get():
            # Usa tema 'clam' per avere più proprietà di stile (troughcolor, arrowcolor, ecc.)
            try:
                style.theme_use("clam")
            except Exception:
                pass
            # Root/backgrounds
            try:
                self.root.configure(bg=bg_dark)
            except Exception:
                pass

            # Base styles
            style.configure("TFrame", background=bg_dark)
            style.configure("TLabelframe", background=bg_dark, foreground=fg_text)
            style.configure("TLabelframe.Label", background=bg_dark, foreground=fg_text)
            style.configure("TLabel", background=bg_dark, foreground=fg_text)
            style.configure("Header.TFrame", background=bg_dark)
            style.configure("Header.TLabel", background=bg_dark, foreground=accent, font=("Segoe UI", 14, "bold"))
            style.configure("SubHeader.TLabel", background=bg_dark, foreground=fg_muted, font=("Segoe UI", 10))
            style.configure("Card.TFrame", background=bg_panel, bordercolor=border, relief="flat")
            style.configure("Card.TLabelframe", background=bg_panel, foreground=fg_text, bordercolor=border, relief="solid")
            style.configure("Card.TLabelframe.Label", background=bg_panel, foreground=fg_muted)
            style.configure("TCheckbutton", background=bg_dark, foreground=fg_text)
            style.configure("TRadiobutton", background=bg_dark, foreground=fg_text)
            style.configure("TCombobox", fieldbackground=bg_input, background=bg_input, foreground=fg_text)
            style.map("TCombobox",
                      fieldbackground=[("readonly", bg_input)],
                      foreground=[("readonly", fg_text)],
                      background=[("readonly", bg_input)])
            style.configure("TEntry", fieldbackground=bg_input, foreground=fg_text, bordercolor=border)
            style.map("TEntry",
                      fieldbackground=[("disabled", bg_input), ("readonly", bg_input)],
                      foreground=[("disabled", fg_muted), ("readonly", fg_text)])

            # Buttons with accent
            style.configure("TButton", background=bg_input, foreground=fg_text, bordercolor=border, relief="flat")
            style.map("TButton",
                      background=[("active", "#30363D"), ("!active", bg_input)],
                      foreground=[("active", "#58A6FF"), ("!active", fg_text)],
                      bordercolor=[("active", accent), ("!active", border)],
                      relief=[("pressed", "sunken"), ("!pressed", "flat")])
            style.configure("Neon.TButton", background=accent, foreground="#FFFFFF", borderwidth=0, focusthickness=2, focuscolor=accent, relief="flat")
            style.map("Neon.TButton",
                      foreground=[("active", "#FFFFFF"), ("!active", "#FFFFFF")],
                      background=[("active", "#58A6FF"), ("!active", accent)],
                      relief=[("pressed", "sunken"), ("!pressed", "flat")])

            # Notebook / Paned / Tree / Scroll
            style.configure("Dark.TNotebook", background=bg_dark)
            style.configure("TNotebook.Tab", background=bg_panel, foreground=fg_muted)
            style.map("TNotebook.Tab", background=[("selected", bg_input)], foreground=[("selected", fg_text)])
            style.configure("Dark.TPanedwindow", background=bg_dark)
            # Scrollbar (clam supporta arrowcolor/troughcolor)
            style.configure("Dark.Vertical.TScrollbar", background=bg_panel, troughcolor=bg_dark,
                            bordercolor=border, arrowcolor=fg_text)
            style.configure("Dark.Horizontal.TScrollbar", background=bg_panel, troughcolor=bg_dark,
                            bordercolor=border, arrowcolor=fg_text)

            # Treeview + Heading scuri
            style.configure("Dark.Treeview",
                            background=bg_panel,
                            fieldbackground=bg_panel,
                            foreground=fg_text,
                            bordercolor=border,
                            rowheight=22)
            style.map("Dark.Treeview",
                      background=[("selected", sel_bg)],
                      foreground=[("selected", accent)])
            style.configure("Treeview.Heading", background=bg_input, foreground=fg_muted, bordercolor=border)
            style.map("Treeview.Heading",
                      background=[("active", "#2D333B")],
                      foreground=[("active", fg_text)])

            # No progressbar in UI (rimosso)

            # Text console
            try:
                self.console.configure(state="normal")
                self.console.configure(
                    bg=bg_panel, fg=fg_text, insertbackground=accent,
                    font=("Consolas", 10), spacing1=0, spacing3=0,
                )
                self.console.tag_configure("neon",      foreground=accent)
                self.console.tag_configure("log_info",  foreground=fg_text)
                self.console.tag_configure("log_debug", foreground=fg_muted)
                self.console.tag_configure("log_warn",  foreground="#D29922")
                self.console.tag_configure("log_err",   foreground="#F85149")
                self.console.tag_configure("log_head",  foreground=accent, font=("Consolas", 10, "bold"))
                self.console.tag_configure("log_raw",   foreground=fg_muted)
                self.console.configure(state="disabled")
            except Exception:
                pass

            # Zebra rows for treeview
            try:
                self.tree.tag_configure("odd", background="#1C2128", foreground=fg_text)
                self.tree.tag_configure("even", background=bg_panel, foreground=fg_text)
                self.tree.tag_configure("hover", background=sel_bg, foreground=accent)
            except Exception:
                pass

            # Applica stile pulsanti per dark
            try:
                if hasattr(self, "btn_open"): self.btn_open.configure(style="Neon.TButton")
                if hasattr(self, "btn_start"): self.btn_start.configure(style="Neon.TButton")
                if hasattr(self, "btn_template"): self.btn_template.configure(style="Neon.TButton")
            except Exception:
                pass
        else:
            # Light professional theme
            l_bg      = "#F6F8FA"
            l_panel   = "#FFFFFF"
            l_input   = "#FFFFFF"
            l_fg      = "#24292F"
            l_muted   = "#57606A"
            l_accent  = "#0969DA"
            l_border  = "#D0D7DE"
            l_sel     = "#DDF4FF"

            try:
                style.theme_use("clam")
            except Exception:
                pass
            try:
                self.root.configure(bg=l_bg)
            except Exception:
                pass

            style.configure("TFrame", background=l_bg)
            style.configure("TLabelframe", background=l_bg, foreground=l_fg, bordercolor=l_border)
            style.configure("TLabelframe.Label", background=l_bg, foreground=l_muted)
            style.configure("TLabel", background=l_bg, foreground=l_fg)
            style.configure("TCheckbutton", background=l_bg, foreground=l_fg)
            style.configure("TRadiobutton", background=l_bg, foreground=l_fg)
            style.configure("TCombobox", fieldbackground=l_input, background=l_input, foreground=l_fg)
            style.map("TCombobox",
                      fieldbackground=[("readonly", l_input)],
                      foreground=[("readonly", l_fg)])
            style.configure("TEntry", fieldbackground=l_input, foreground=l_fg, bordercolor=l_border)
            style.configure("TButton", background=l_bg, foreground=l_fg, bordercolor=l_border, relief="flat")
            style.map("TButton",
                      background=[("active", "#EAEEF2"), ("!active", l_bg)],
                      relief=[("pressed", "sunken"), ("!pressed", "flat")])
            style.configure("Neon.TButton", background=l_accent, foreground="#FFFFFF", borderwidth=0, relief="flat")
            style.map("Neon.TButton",
                      background=[("active", "#0860CA"), ("!active", l_accent)],
                      foreground=[("active", "#FFFFFF"), ("!active", "#FFFFFF")],
                      relief=[("pressed", "sunken"), ("!pressed", "flat")])
            style.configure("Dark.TNotebook", background=l_bg)
            style.configure("TNotebook.Tab", background=l_bg, foreground=l_muted)
            style.map("TNotebook.Tab", background=[("selected", l_panel)], foreground=[("selected", l_fg)])
            style.configure("Dark.TPanedwindow", background=l_bg)
            style.configure("Dark.Vertical.TScrollbar", background=l_bg, troughcolor=l_bg, bordercolor=l_border, arrowcolor=l_muted)
            style.configure("Dark.Horizontal.TScrollbar", background=l_bg, troughcolor=l_bg, bordercolor=l_border, arrowcolor=l_muted)
            style.configure("Dark.Treeview", background=l_panel, fieldbackground=l_panel, foreground=l_fg, bordercolor=l_border, rowheight=22)
            style.map("Dark.Treeview", background=[("selected", l_sel)], foreground=[("selected", l_accent)])
            style.configure("Treeview.Heading", background=l_bg, foreground=l_muted, bordercolor=l_border)
            style.map("Treeview.Heading", background=[("active", "#EAEEF2")], foreground=[("active", l_fg)])
            style.configure("Header.TFrame", background=l_bg)
            style.configure("Header.TLabel", background=l_bg, foreground=l_accent, font=("Segoe UI", 14, "bold"))
            style.configure("SubHeader.TLabel", background=l_bg, foreground=l_muted, font=("Segoe UI", 10))
            style.configure("Card.TFrame", background=l_panel, bordercolor=l_border, relief="flat")
            style.configure("Card.TLabelframe", background=l_panel, foreground=l_fg, bordercolor=l_border, relief="solid")
            style.configure("Card.TLabelframe.Label", background=l_panel, foreground=l_muted)
            try:
                self.console.configure(state="normal")
                self.console.configure(
                    bg=l_panel, fg=l_fg, insertbackground=l_accent,
                    font=("Consolas", 10), spacing1=0, spacing3=0,
                )
                self.console.tag_configure("neon",      foreground=l_accent)
                self.console.tag_configure("log_info",  foreground=l_fg)
                self.console.tag_configure("log_debug", foreground=l_muted)
                self.console.tag_configure("log_warn",  foreground="#9A6700")
                self.console.tag_configure("log_err",   foreground="#CF222E")
                self.console.tag_configure("log_head",  foreground=l_accent, font=("Consolas", 10, "bold"))
                self.console.tag_configure("log_raw",   foreground=l_muted)
                self.console.configure(state="disabled")
            except Exception:
                pass
            try:
                self.tree.tag_configure("odd", background="#F6F8FA", foreground=l_fg)
                self.tree.tag_configure("even", background=l_panel, foreground=l_fg)
                self.tree.tag_configure("hover", background=l_sel, foreground=l_accent)
            except Exception:
                pass
            # Applica stile pulsanti per light
            try:
                if hasattr(self, "btn_open"): self.btn_open.configure(style="TButton")
                if hasattr(self, "btn_start"): self.btn_start.configure(style="Neon.TButton")
                if hasattr(self, "btn_template"): self.btn_template.configure(style="TButton")
            except Exception:
                pass

    def _build_sheet_options(self, parent):
        """Costruisce le opzioni di selezione lastra."""
        row1 = ttk.Frame(parent)
        row1.pack(fill=tk.X, padx=4, pady=2)
        ttk.Radiobutton(
            row1, text="LASTRA INOX 1250×3000",
            variable=self.sheet_choice, value="inox",
            command=self._on_sheet_choice
        ).pack(side=tk.LEFT)

        row2 = ttk.Frame(parent)
        row2.pack(fill=tk.X, padx=4, pady=2)
        ttk.Radiobutton(
            row2, text="LASTRA HPL 1300×3050",
            variable=self.sheet_choice, value="hpl",
            command=self._on_sheet_choice
        ).pack(side=tk.LEFT)

        row3 = ttk.Frame(parent)
        row3.pack(fill=tk.X, padx=4, pady=2)
        ttk.Radiobutton(
            row3, text="LASTRA CORIAN 930×3000",
            variable=self.sheet_choice, value="corian",
            command=self._on_sheet_choice
        ).pack(side=tk.LEFT)

        row3b = ttk.Frame(parent)
        row3b.pack(fill=tk.X, padx=4, pady=2)
        ttk.Radiobutton(
            row3b, text="LASTRA CORIAN + GRECATO 1200×2945",
            variable=self.sheet_choice, value="corian_grecato",
            command=self._on_sheet_choice
        ).pack(side=tk.LEFT)

        row4 = ttk.Frame(parent)
        row4.pack(fill=tk.X, padx=4, pady=2)
        ttk.Radiobutton(
            row4, text="LASTRA CUSTOM",
            variable=self.sheet_choice, value="custom",
            command=self._on_sheet_choice
        ).pack(side=tk.LEFT)
        ttk.Label(row4, text="W:").pack(side=tk.LEFT, padx=(12,3))
        self.ent_w = ttk.Entry(row4, width=8, textvariable=self.sheet_w, state="disabled")
        self.ent_w.pack(side=tk.LEFT)
        ttk.Label(row4, text="H:").pack(side=tk.LEFT, padx=(12,3))
        self.ent_h = ttk.Entry(row4, width=8, textvariable=self.sheet_h, state="disabled")
        self.ent_h.pack(side=tk.LEFT)
        
        self._on_sheet_choice()

    # =====================
    # Output XLSX Viewer
    # =====================
    def _build_output_viewer(self, parent):
        bar = ttk.Frame(parent, style="Card.TFrame")
        bar.pack(fill=tk.X, padx=6, pady=(6,0))
        ttk.Label(bar, text="File XLSX:").pack(side=tk.LEFT)
        self.xlsx_path_var = tk.StringVar(value="")
        self.ent_xlsx = ttk.Entry(bar, textvariable=self.xlsx_path_var, width=60)
        self.ent_xlsx.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=6)
        self.btn_browse_xlsx = ttk.Button(bar, text="Apri…", command=self._browse_xlsx)
        self.btn_browse_xlsx.pack(side=tk.LEFT)
        self.btn_load_xlsx = ttk.Button(bar, text="Carica", command=self._load_xlsx)
        self.btn_load_xlsx.pack(side=tk.LEFT, padx=(6,0))

        # Notebook per fogli
        self.nb_xlsx = ttk.Notebook(parent, style="Dark.TNotebook")
        self.nb_xlsx.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self._xlsx_tabs = {}

    def _browse_xlsx(self):
        path = filedialog.askopenfilename(
            title="Seleziona workbook XLSX",
            filetypes=[("Excel", "*.xlsx"), ("Tutti", "*.*")]
        )
        if path:
            self.xlsx_path_var.set(path)

    def _load_xlsx(self):
        path = self.xlsx_path_var.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showwarning("XLSX mancante", "Seleziona un file XLSX valido.")
            return
        try:
            xls = pd.ExcelFile(path)
        except Exception as e:
            messagebox.showerror("Errore apertura XLSX", str(e))
            return

        # Ricrea i tab in base ai fogli disponibili (mostra fino a 4 principali)
        for name in list(self._xlsx_tabs.keys()):
            tab, tree = self._xlsx_tabs[name]
            try:
                idx = self.nb_xlsx.index(tab)
                self.nb_xlsx.forget(idx)
            except Exception:
                pass
        self._xlsx_tabs.clear()

        wanted = ["LAYOUT", "SHEETS", "MACHINE", "PACKAGING"]
        names = [n for n in wanted if n in xls.sheet_names] or xls.sheet_names[:4]

        for name in names:
            try:
                df = xls.parse(name)
            except Exception:
                continue
            tab = ttk.Frame(self.nb_xlsx, style="Card.TFrame")
            self.nb_xlsx.add(tab, text=name)
            tree = ttk.Treeview(tab, columns=list(df.columns), show="headings", style="Dark.Treeview")
            tree.pack(fill=tk.BOTH, expand=True)
            # intestazioni
            for c in df.columns:
                tree.heading(c, text=str(c))
                tree.column(c, width=120, stretch=True)
            # righe
            for i, row in df.iterrows():
                tag = "odd" if (i % 2) else "even"
                vals = [row.get(c, "") for c in df.columns]
                tree.insert("", tk.END, values=vals, tags=(tag,))
            # mousewheel
            tree.bind('<MouseWheel>', lambda e, t=tree: self._tv_wheel(e, t))
            tree.bind('<Shift-MouseWheel>', lambda e, t=tree: self._tv_wheel_h(e, t))
            self._xlsx_tabs[name] = (tab, tree)

    def _tv_wheel(self, event, tree):
        delta = int(-1*(event.delta/120)) if hasattr(event, 'delta') else 0
        try:
            tree.yview_scroll(delta, 'units')
        except Exception:
            pass
        return "break"

    def _tv_wheel_h(self, event, tree):
        delta = int(-1*(event.delta/120)) if hasattr(event, 'delta') else 0
        try:
            tree.xview_scroll(delta, 'units')
        except Exception:
            pass
        return "break"

    def _on_sheet_choice(self):
        """Abilita/disabilita campi custom."""
        if self.sheet_choice.get() == "custom":
            self.ent_w.configure(state="normal")
            self.ent_h.configure(state="normal")
        else:
            self.ent_w.configure(state="disabled")
            self.ent_h.configure(state="disabled")

    def _on_drop(self, event):
        """Handler drag & drop."""
        path = event.data.strip("{}")
        self.selected_file.set(path)
        self.load_preview(path)

    def open_file(self):
        """Apri file dialog."""
        path = filedialog.askopenfilename(
            title="Seleziona file pannelli",
            filetypes=[("Excel", "*.xlsx;*.xls"), ("Tutti", "*.*")]
        )
        if not path:
            return
        self.selected_file.set(path)
        self.load_preview(path)

    def download_template(self):
        """Salva un file XLSX template con le sole intestazioni delle colonne."""
        path = filedialog.asksaveasfilename(
            title="Salva template XLSX",
            defaultextension=".xlsx",
            initialfile="template_pannelli.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Pannelli"
            headers = ["id", "number", "width", "height", "qty"]
            ws.append(headers)

            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            fill = PatternFill("solid", fgColor="F2F2F2")
            for col_idx, name in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = fill
                cell.border = border
                ws.column_dimensions[cell.column_letter].width = max(12, len(name) + 6)

            wb.save(path)
            messagebox.showinfo("Template salvato", f"Template creato:\n{path}")
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile creare il template:\n{e}")

    def load_preview(self, path):
        """Carica anteprima file nella treeview."""
        try:
            if path.lower().endswith(".csv"):
                df = pd.read_csv(path)
            else:
                df = pd.read_excel(path)
        except Exception as e:
            messagebox.showerror("Errore apertura", str(e))
            return
        
        dfp = df.head(500).copy()
        
        for c in self.tree["columns"]:
            self.tree.heading(c, text="")
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = list(dfp.columns)
        
        for c in dfp.columns:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=120, stretch=True)
        
        for i, (_, row) in enumerate(dfp.iterrows()):
            tag = "odd" if (i % 2) else "even"
            self.tree.insert("", tk.END, values=[row.get(c, "") for c in dfp.columns], tags=(tag,))
        try:
            self._init_tree_hover()
        except Exception:
            pass

    def _init_tree_hover(self):
        """Hover row highlight via dynamic tag swap (ttk.Treeview compatibile)."""
        self._tree_hover = None

        def _swap_tag(rowid, add):
            try:
                tags = list(self.tree.item(rowid, "tags") or ())
            except Exception:
                return
            if add:
                if "hover" not in tags:
                    tags.append("hover")
            else:
                tags = [t for t in tags if t != "hover"]
            try:
                self.tree.item(rowid, tags=tags)
            except Exception:
                pass

        def on_motion(event):
            try:
                rowid = self.tree.identify_row(event.y)
            except Exception:
                return
            if rowid != self._tree_hover:
                if self._tree_hover:
                    _swap_tag(self._tree_hover, add=False)
                self._tree_hover = rowid
                if rowid:
                    _swap_tag(rowid, add=True)

        def on_leave(_):
            if getattr(self, "_tree_hover", None):
                _swap_tag(self._tree_hover, add=False)
                self._tree_hover = None

        self.tree.bind("<Motion>", on_motion)
        self.tree.bind("<Leave>", on_leave)

    def ask_output_name(self):
        """Richiede nome cartella output."""
        if not self.selected_file.get():
            messagebox.showwarning("Seleziona file", "Seleziona prima un file di input.")
            return
        
        input_path = self.selected_file.get()
        input_basename = os.path.splitext(os.path.basename(input_path))[0]
        
        choice = self.sheet_choice.get()
        if choice == "inox":
            suffix = "_INOX"
        elif choice == "hpl":
            suffix = "_HPL"
        elif choice == "corian":
            suffix = "_CORIAN"
        elif choice == "corian_grecato":
            suffix = "_CORIAN+GRECATO"
        else:
            suffix = ""
        
        default_name = f"{input_basename}{suffix}"
        out_name = simpledialog.askstring("Nome cartella output",
                                          "Inserisci il nome/prefisso della cartella di output:",
                                          initialvalue=default_name)
        if not out_name:
            return
        self.out_prefix = out_name.strip()
        self.start_run()

    def start_run(self):
        """Avvia il calcolo in thread separato."""
        choice = self.sheet_choice.get()
        if choice == "none":
            messagebox.showwarning("Seleziona dimensioni", "Seleziona una lastra.")
            return

        # Valida kerf
        try:
            k = float(self.kerf.get().strip())
            if k < 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Dati non validi", "Kerf deve essere un numero ≥ 0.")
            return

        # Determina W/H
        if choice == "inox":
            tag, w, h = "LASTRA_INOX_1250x3000", 1250.0, 3000.0
        elif choice == "hpl":
            tag, w, h = "LASTRA_HPL_1300x3050", 1300.0, 3050.0
        elif choice == "corian":
            tag, w, h = "LASTRA_CORIAN_930x3000", 930.0, 3000.0
        elif choice == "corian_grecato":
            tag, w, h = "LASTRA_CORIAN_GRECATO_1200x2945", 1200.0, 2945.0
        elif choice == "custom":
            try:
                w = float(self.sheet_w.get().strip())
                h = float(self.sheet_h.get().strip())
                if w <= 0 or h <= 0:
                    raise ValueError
            except Exception:
                messagebox.showerror("Dati non validi", "Inserisci W/H valide per CUSTOM.")
                return
            tag = f"LASTRA_CUSTOM_{int(w)}x{int(h)}"
        else:
            messagebox.showerror("Errore", f"Scelta sconosciuta: {choice}")
            return
        
        inox_adjust = (choice == "inox")

        # Materiale
        if choice == "inox":
            mat = "inox"
        elif choice == "hpl":
            mat = "hpl"
        elif choice == "corian":
            mat = "corian"
        elif choice == "corian_grecato":
            mat = "corian_grecato"
        else:
            mat = "hpl"

        # Peso max cassa
        try:
            maxkg = float(self.max_crate_kg.get().strip())
            if maxkg <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Dati non validi", "Peso massimo cassa deve essere > 0.")
            return

        # Path file
        import pathlib
        sel_path = self.selected_file.get().strip()
        if not sel_path:
            messagebox.showwarning("Seleziona file", "Seleziona prima un file di input.")
            return

        # Kwargs
        base_kwargs = dict(
            file=pathlib.Path(sel_path),
            mode="heuristic",
            kerf=k,
            no_rotate=self.no_rotate.get(),
            visualise=self.visualise.get(),
            group_run=self.group_run.get(),
            id_order=self.id_order.get(),
            timeout=300,
            material=mat,
            max_crate_kg=maxkg,
        )

        # La policy GUI fa già un filtraggio user-friendly migliore: disattiva
        # il filtro globale per ricevere tutti i messaggi e poterli trasformare.
        try:
            CONCISE_FILTER.enabled = False
        except Exception:
            pass
        # Toggle "Log sintetico": attivo = trasformazione user-friendly,
        # disattivo = passa-tutto (modalità sviluppatore).
        try:
            self._handler.policy = self._policy if self.concise_log.get() else None
        except Exception:
            pass

        sub_out = self.out_prefix

        def _runner():
            try:
                # Import qui per evitare dipendenze circolari
                from main import run_with_namespace

                # Reset stato della policy ad ogni run
                try:
                    self._policy.reset()
                except Exception:
                    pass

                args = _ArgsNamespace(**base_kwargs, sheet=[w, h], out=sub_out, inox_adjust=inox_adjust)
                self.write_section(f"▶  {tag}  •  W={w:g} H={h:g}  →  {sub_out}")
                run_with_namespace(args)
                self._q.put(("LOG", f"\n└─ ✅ Completato {tag}\n", "HEADER"))
                self._q.put(("LOG", "🎯 Job completato.\n", "INFO"))

            except Exception as e:
                import traceback
                self._q.put(("LOG", f"❌ Errore: {e}\n", "ERROR"))
                self._q.put(("RAW", traceback.format_exc(), "RAW"))
                

        t = threading.Thread(target=_runner, daemon=True)
        t.start()

    def _setup_logging(self):
        """Setup logger e redirect stdout/stderr."""
        self._policy = UserFriendlyPolicy()
        self._handler = TextLoggerHandler(self._write_log, policy=self._policy)
        # nessun setFormatter: l'handler costruisce il formato da sé
        logging.getLogger().addHandler(self._handler)

        sys.stdout = StreamToText(self._write_raw)
        sys.stderr = StreamToText(self._write_raw)

        # Stato per collassare righe vuote consecutive nello stdout grezzo
        self._raw_buf = ""
        self._last_blank = True  # comincia "soffocando" eventuali blank iniziali

    def _write_log(self, text, level="INFO"):
        """Thread-safe write log con livello."""
        self._q.put(("LOG", text, level))

    def _write_raw(self, text):
        """Thread-safe write raw (stdout/stderr).

        Bufferizza per riga e collassa newline consecutivi (riduce gli spazi
        verticali generati dai print con \n extra).
        """
        if not text:
            return
        self._raw_buf += text
        while "\n" in self._raw_buf:
            line, self._raw_buf = self._raw_buf.split("\n", 1)
            stripped = line.strip()
            if not stripped:
                # blank: ignora se già blank o se policy attiva (riduce vuoti)
                continue
            pairs = self._policy.transform_raw(line) if getattr(self, "_policy", None) else [(line, "RAW")]
            if not pairs:
                continue
            ts = _time.strftime("%H:%M:%S")
            for msg, lvl in pairs:
                if lvl == "RAW":
                    # mantiene aspetto raw, ma indenta
                    self._q.put(("RAW", "          " + msg + "\n", "RAW"))
                else:
                    glyph = LEVEL_GLYPH.get(lvl, " ")
                    self._q.put(("LOG", f"{ts}  {glyph}  {msg}\n", lvl))
            self._last_blank = False

    def write_section(self, title):
        """Scrive un'intestazione di sezione nella console."""
        bar = "─" * max(8, 60 - len(title))
        self._q.put(("LOG", f"\n┌─ {title} {bar}\n", "HEADER"))

    def _drain(self):
        """Svuota coda e aggiorna Text widget con tag per livello."""
        tag_map = {
            "DEBUG":   "log_debug",
            "INFO":    "log_info",
            "WARNING": "log_warn",
            "ERROR":   "log_err",
            "CRITICAL":"log_err",
            "HEADER":  "log_head",
            "RAW":     "log_raw",
        }
        try:
            while True:
                item = self._q.get_nowait()
                if not item:
                    continue
                if len(item) == 2:
                    typ, payload = item
                    level = "INFO"
                else:
                    typ, payload, level = item
                tag = tag_map.get(level, "log_info")
                self.console.configure(state="normal")
                self.console.insert(tk.END, payload, (tag,))
                if not payload.endswith("\n"):
                    self.console.insert(tk.END, "\n", (tag,))
                self.console.see(tk.END)
                self.console.configure(state="disabled")
        except queue.Empty:
            pass
        self.root.after(80, self._drain)


# ============================================================================
# MAIN GUI
# ============================================================================

def gui_main():
    """Entry point per GUI."""
    if not _HAS_TK:
        print("Tkinter non disponibile in questo ambiente.")
        sys.exit(1)
    
    if _HAS_DND:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
    
    # Tema
    try:
        style = ttk.Style(root)
        if "vista" in style.theme_names():
            style.theme_use("vista")
    except Exception:
        pass
    
    try:
        _ico = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "print_cubes_scheme_modular_color_edit_products_miscellaneous_icon_251257.ico")
        if os.path.isfile(_ico):
            root.iconbitmap(_ico)
    except Exception:
        pass

    app = NestGUI(root)
    root.geometry("1200x720")
    root.mainloop()